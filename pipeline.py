"""
Full OCR pipeline for historical Indian census tables.

Strategies:
  1. Full-image extraction (baseline)
  2. Row-by-row cropped extraction (zoom)
  3. Multi-pass voting (3 passes per model)
  4. Cross-model ensemble
  5. Constraint enforcement (M+F=Persons, sums)
  6. Multi-Enhancement MoE (4 variants × 3 models = 12 readings, digit-level ensemble)

Usage:
    python3 pipeline.py                          # run all strategies on all tests
    python3 pipeline.py --test 0                 # specific test
    python3 pipeline.py --strategy full          # just baseline
    python3 pipeline.py --strategy multipass     # 3-pass voting
    python3 pipeline.py --strategy ensemble      # cross-model + constraints
    python3 pipeline.py --strategy cropped       # row-cropped extraction
    python3 pipeline.py --strategy moe           # multi-enhancement MoE pipeline
    python3 pipeline.py --strategy all           # everything combined
"""

import os
import sys
import json
import base64
import time
import tempfile
import argparse
import logging
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

logging.basicConfig(level=logging.INFO, format="%(name)s: %(message)s")

load_dotenv()

PROJECT_ROOT = Path(__file__).parent
AGE_TABLES = PROJECT_ROOT / "age_tables"
DATA_DIR = PROJECT_ROOT / "Data"
RESULTS_DIR = PROJECT_ROOT / "results"
RESULTS_DIR.mkdir(exist_ok=True)

sys.path.insert(0, str(PROJECT_ROOT))
from image_processing import crop_rows, crop_vertical_sections, crop_region
from ensemble import (majority_vote, cross_model_ensemble, enforce_constraints,
                      digit_level_vote, digit_level_ensemble, constraint_ensemble,
                      cross_group_reconcile)


# ─── Utilities ───────────────────────────────────────────────────────────

def fix_excel_age(age):
    if hasattr(age, 'strftime'):
        return f"{age.month}-{age.day}"
    return str(age)

def normalize_age(age_str):
    import re
    s = str(age_str).replace("\u2013", "-").replace("\u2014", "-").replace(" ", "").lower()
    # Normalize "&" to "and" so "60 & over" matches "60 and over"
    s = s.replace("&", "and")
    # Collapse double-hyphens (OCR artifact: "15--19" -> "15-19")
    s = re.sub(r'-{2,}', '-', s)
    # Strip trailing periods (OCR artifact: "40-44." -> "40-44")
    s = s.rstrip(".")
    # Strip "total" prefix (e.g. "Total 0-5" -> "0-5") but preserve bare "total"
    stripped = re.sub(r'^total\s*', '', s).strip()
    return stripped if stripped else s.strip()

def encode_image(image_path):
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def encode_pil_image(pil_img):
    import io
    buf = io.BytesIO()
    pil_img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")

def parse_json_response(text):
    import re as _re
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        text = "\n".join(lines)
    text = _re.sub(r'(?<=\d)_(?=\d)', '', text)
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        start = text.find("[")
        end = text.rfind("]") + 1
        if start >= 0 and end > start:
            try:
                parsed = json.loads(text[start:end])
            except json.JSONDecodeError:
                return None
        else:
            return None
    # Unwrap JSON object wrapping an array (e.g. {"rows": [...]})
    if isinstance(parsed, dict):
        lists = [v for v in parsed.values() if isinstance(v, list)]
        if len(lists) == 1:
            return lists[0]
        # If multiple list values, pick the longest
        if lists:
            return max(lists, key=len)
        return None
    return parsed


# ─── Ground Truth ────────────────────────────────────────────────────────

def load_gt_simple(xlsx_name, sheet_name, district_filter=None):
    wb = openpyxl.load_workbook(DATA_DIR / xlsx_name, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        district = row[1]
        if district is None:
            continue
        if district_filter and district != district_filter:
            continue
        age = fix_excel_age(row[2])
        persons = row[3]
        males = row[4]
        females = row[5]
        if persons is None:
            continue
        rows.append({
            "age": age,
            "persons": int(persons),
            "males": int(float(males)) if males else 0,
            "females": int(float(females)) if females else 0,
        })
    return rows


def load_gt_coorg_1901():
    """Load Coorg 1901 GT, aggregating individual-year rows 0-1..4-5 into a 0-5 summary."""
    raw = load_gt_simple("Coorg.xlsx", "Coorg_1901", "Coorg")
    # Aggregate 0-1 through 4-5 into a single 0-5 row
    individual = []
    summary = []
    for row in raw:
        age = normalize_age(row["age"])
        # Check if it's an individual year group within 0-5
        if age in ("0-1", "1-2", "2-3", "3-4", "4-5"):
            individual.append(row)
        else:
            summary.append(row)
    if individual:
        agg = {
            "age": "0-5",
            "persons": sum(r["persons"] for r in individual),
            "males": sum(r["males"] for r in individual),
            "females": sum(r["females"] for r in individual),
        }
        summary.insert(0, agg)
    return summary


# GT corrections: override known errors in Excel files
GT_CORRECTIONS = {
    "travancore_east_1901": {
        "20-25": {"persons": 111950, "males": 53002},
        "30-35": {"persons": 95124, "females": 45867},
    },
    "coorg_1901": {
        # User adjudicated: model reads the scan correctly in both cases
        "0-5": {"persons": 218, "males": 95},   # Excel individual years sum to 219/96
        "50-55": {"persons": 67, "males": 32},   # Excel has 78/43 — confirmed wrong
    },
    "hyderabad_state_1901": {
        "5-10": {"persons": 1452854, "males": 755963, "females": 696891},
        "10-15": {"persons": 1350321, "males": 739509},
        "15-20": {"persons": 852389, "males": 428409},
        "20-25": {"persons": 894027, "males": 405798, "females": 488229},
        "25-30": {"persons": 1051503, "males": 523613},
        "35-40": {"persons": 628589, "females": 285976},
        "40-45": {"persons": 818675, "females": 392922},
        "45-50": {"persons": 357058, "females": 156244},
        "60andover": {"persons": 570951, "males": 271002, "females": 299949},
    },
}


def apply_gt_corrections(gt_rows, test_id):
    """Apply known corrections to ground truth data."""
    corrections = GT_CORRECTIONS.get(test_id)
    if not corrections:
        return gt_rows
    corrected = []
    for row in gt_rows:
        row = dict(row)
        age_key = normalize_age(row.get("age", ""))
        if age_key in corrections:
            for col, val in corrections[age_key].items():
                row[col] = val
        corrected.append(row)
    return corrected


TEST_CASES = [
    {
        "id": "travancore_east_1901",
        "name": "Travancore Eastern Division 1901",
        "image_path": AGE_TABLES / "Travancore/1901/Eastern_division_age_1901.png",
        "ground_truth_loader": lambda: load_gt_simple(
            "Travancore.xlsx", "Travancore_1901", "Eastern Division"),
        "preprocessing": {"crop": None},
    },
    {
        "id": "hyderabad_state_1901",
        "name": "Hyderabad State Summary 1901",
        "image_path": AGE_TABLES / "Hyderabad/Hyderabad_state_summary_age_1901.png",
        "ground_truth_loader": lambda: load_gt_simple(
            "Hyderabad.xlsx", "Hyderabad_1901", "Hyderabad"),
        "preprocessing": {"crop": {"right_frac": 0.33}},
    },
    {
        "id": "coorg_1901",
        "name": "Coorg 1901",
        "image_path": AGE_TABLES / "Coorg/1901/Coorg_age_1901.png",
        "ground_truth_loader": load_gt_coorg_1901,
        "preprocessing": {"crop": {"right_frac": 0.65}},
        "prompt": "subtotals",  # Use PROMPT_FULL_WITH_SUBTOTALS
        "constraint_groups": {
            # Subtotal rows and their components (normalized age keys)
            # _normalize_age("Total 0-5") -> "0-5", "Total 0-15" -> "0-15", etc.
            "0-5": ["0-1", "1-2", "2-3", "3-4", "4-5"],
            "0-15": ["0-5", "5-10", "10-15"],
            "15-40": ["15-20", "20-25", "25-30", "30-35", "35-40"],
            "40-60": ["40-45", "45-50", "50-55", "55-60"],
            "total": ["0-15", "15-40", "40-60", "60andover"],
        },
        "known_totals": {
            # For per-1000 tables, the grand total is always 1000
            "total": {"males": 1000, "females": 1000},
        },
    },
]


# ─── API Callers ─────────────────────────────────────────────────────────

def call_openai(b64_image, prompt, mime="image/png"):
    from openai import OpenAI
    client = OpenAI()
    response = client.chat.completions.create(
        model="gpt-4.1",
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {
                    "url": f"data:{mime};base64,{b64_image}",
                    "detail": "high"
                }},
            ],
        }],
        max_tokens=8192,
        temperature=0.1,  # slight temp for diversity in multi-pass
        # Note: response_format=json_object not used — forces valid JSON but
        # degrades data quality; parse failures correctly filter garbage readings
    )
    return response.choices[0].message.content


# ---------------------------------------------------------------------------
# Gemini client cache + rate limiting
# ---------------------------------------------------------------------------
_gemini_client = None
_gemini_client_key = None
_last_gemini_call = 0.0

def _get_gemini_client():
    """Return a cached Gemini client, creating one if needed."""
    global _gemini_client, _gemini_client_key
    from google import genai
    api_key = os.environ["GEMINI_API_KEY"]
    if _gemini_client is None or _gemini_client_key != api_key:
        _gemini_client = genai.Client(
            api_key=api_key,
            http_options=genai.types.HttpOptions(timeout=120_000),
        )
        _gemini_client_key = api_key
    return _gemini_client


def call_gemini(b64_image, prompt, mime="image/png"):
    global _last_gemini_call
    import time as _time
    from google import genai

    # Rate limiting: enforce minimum interval between calls.
    # Set GEMINI_MIN_INTERVAL=1.0 (seconds) to throttle. Default 0 = no limit.
    min_interval = float(os.environ.get("GEMINI_MIN_INTERVAL", "0"))
    if min_interval > 0 and _last_gemini_call > 0:
        elapsed = _time.time() - _last_gemini_call
        if elapsed < min_interval:
            _time.sleep(min_interval - elapsed)

    client = _get_gemini_client()
    primary_model = os.environ.get("GEMINI_MODEL", "gemini-3.1-pro-preview")
    fallback_model = "gemini-2.5-pro"
    allow_fallback = os.environ.get("GEMINI_ALLOW_FALLBACK", "0") == "1"
    contents = [
        prompt,
        genai.types.Part.from_bytes(data=base64.b64decode(b64_image), mime_type=mime),
    ]
    config = genai.types.GenerateContentConfig(temperature=0.1)

    # Retry primary model with exponential backoff
    max_attempts = 5
    for attempt in range(max_attempts):
        _last_gemini_call = _time.time()
        try:
            response = client.models.generate_content(
                model=primary_model, contents=contents, config=config,
            )
            return response.text
        except Exception as e:
            err_str = str(e)
            is_server_error = any(code in err_str for code in ['500', '503', '504', '429', 'UNAVAILABLE', 'RESOURCE_EXHAUSTED', 'ServerError', 'INTERNAL'])
            if is_server_error and attempt < max_attempts - 1:
                wait = 15 * (2 ** attempt)  # 15, 30, 60, 120
                print(f"    Gemini {primary_model} error (attempt {attempt+1}/{max_attempts}): {err_str[:80]}")
                print(f"    Retrying in {wait}s...")
                _time.sleep(wait)
            elif is_server_error and allow_fallback and primary_model != fallback_model:
                print(f"    WARNING: {primary_model} failed {max_attempts} times, "
                      f"falling back to {fallback_model}")
                break  # Fall through to fallback
            else:
                raise

    # Fallback only when explicitly allowed via GEMINI_ALLOW_FALLBACK=1
    if allow_fallback and primary_model != fallback_model:
        for attempt in range(3):
            _last_gemini_call = _time.time()
            try:
                response = client.models.generate_content(
                    model=fallback_model, contents=contents, config=config,
                )
                return response.text
            except Exception as e:
                err_str = str(e)
                is_server_error = any(code in err_str for code in ['500', '503', '504', '429', 'UNAVAILABLE', 'RESOURCE_EXHAUSTED', 'ServerError', 'INTERNAL'])
                if is_server_error and attempt < 2:
                    wait = 10 * (2 ** attempt)
                    print(f"    Gemini {fallback_model} error (attempt {attempt+1}/3): {err_str[:80]}")
                    print(f"    Retrying in {wait}s...")
                    _time.sleep(wait)
                else:
                    raise
    raise Exception(f"Gemini {primary_model} failed after {max_attempts} attempts")


def call_claude(b64_image, prompt, mime="image/png"):
    import anthropic
    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-opus-4-20250514",
        max_tokens=8192,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {
                    "type": "base64",
                    "media_type": mime,
                    "data": b64_image,
                }},
                {"type": "text", "text": prompt},
            ],
        }],
    )
    return response.content[0].text


MODELS = {
    "openai": call_openai,
    "gemini": call_gemini,
    "claude": call_claude,
}


# ─── Prompts ─────────────────────────────────────────────────────────────

PROMPT_FULL = """You are an expert at reading historical census tables. This is a scanned page from an Indian census report.

Extract the POPULATION totals (Persons, Males, Females columns) for each SUMMARY age group.
Summary age groups: 0-5, 5-10, 10-15, 15-20, 20-25, 25-30, 30-35, 35-40, 40-45, 45-50, 50-55, 55-60, 60 and over.
Also include the grand Total row if visible.

CRITICAL: Read each digit extremely carefully. Common OCR errors on old typefaces:
- 6 vs 5, 3 vs 8, 0 vs 9, 1 vs 7
- Verify by checking: Persons should equal Males + Females for each row
- If a digit is ambiguous, use the M+F=Persons constraint to determine the correct reading

Return ONLY a JSON array. Each object: {"age": "...", "persons": N, "males": N, "females": N}
All values are integers, no commas. Return ONLY valid JSON."""

PROMPT_FULL_WITH_SUBTOTALS = """You are an expert at reading historical census tables. This is a scanned page from an Indian census report.

This table shows Males and Females columns (there may be no Persons column — if so, compute Persons = Males + Females).

Extract EVERY row visible in the table, including:
- Individual age groups (0-1, 1-2, etc. if present)
- Summary age groups (0-5, 5-10, 10-15, 15-20, 20-25, 25-30, 30-35, 35-40, 40-45, 45-50, 50-55, 55-60, 60 and over)
- ALL subtotal rows (Total 0-5, Total 0-15, Total 15-40, Total 40-60, etc.)
- The grand Total row

CRITICAL: Read each digit extremely carefully. In this typeface:
- The digit 3 has angular notches/points at its curves; 8 has smooth continuous curves. These look very similar — examine closely.
- Also watch for: 6 vs 5, 0 vs 9, 1 vs 7
- Verify by checking: Persons should equal Males + Females for each row
- Subtotal rows should equal the sum of their component rows

Return ONLY a JSON array. Each object: {"age": "...", "persons": N, "males": N, "females": N}
For subtotal rows, use the label as-is (e.g. "Total 0-5", "Total 15-40", "Total").
All values are integers, no commas. Return ONLY valid JSON."""

PROMPT_CROPPED_ROW = """You are an expert at reading historical census tables. This image shows a CROPPED ROW from a census table.

The table has columns: Age, Persons, Males, Females (and possibly more columns to the right like Unmarried, Married, Widowed).

Extract ONLY the POPULATION section values (the first set of Persons, Males, Females columns after the Age column).
Read each digit very carefully. For old typefaces: 6 can look like 5, 3 like 8, 0 like 9.

If you can identify the age group, include it. If you see multiple rows, extract all of them.

Return ONLY a JSON array: [{"age": "...", "persons": N, "males": N, "females": N}]
All values are integers. Return ONLY valid JSON."""

PROMPT_VERIFY = """You are verifying OCR results from a historical census table. Here is the original image.

I have a preliminary reading of the data. For each row, please verify the numbers by reading the image very carefully.
Check especially:
1. Does Persons = Males + Females?
2. Are there any digits that look ambiguous (6/5, 3/8, 0/9)?

Previous reading:
{previous_reading}

Return the CORRECTED data as a JSON array: [{{"age": "...", "persons": N, "males": N, "females": N}}]
Fix any errors you find. All values are integers. Return ONLY valid JSON."""


# ─── Scoring ─────────────────────────────────────────────────────────────

def score(predicted, ground_truth):
    if predicted is None:
        return {"exact_match_rate": 0, "error": "null prediction"}

    gt_by_age = {}
    for row in ground_truth:
        gt_by_age[normalize_age(row["age"])] = row

    total = 0
    exact = 0
    abs_errors = []
    details = []

    matched_keys = set()
    for pred_row in predicted:
        pred_age = normalize_age(pred_row.get("age", ""))
        if not pred_age:
            continue
        # Exact match only (normalize_age handles all formatting differences)
        if pred_age not in gt_by_age or pred_age in matched_keys:
            continue
        gt_row = gt_by_age[pred_age]

        matched_keys.add(normalize_age(gt_row["age"]))
        for col in ["persons", "males", "females"]:
            gt_val = gt_row.get(col)
            if gt_val is None:
                continue
            pred_val = pred_row.get(col)
            if pred_val is None:
                total += 1
                abs_errors.append(abs(gt_val))
                continue
            try:
                pred_val = int(str(pred_val).replace(",", "").replace(" ", ""))
            except (ValueError, TypeError):
                total += 1
                abs_errors.append(abs(gt_val))
                continue
            total += 1
            if pred_val == gt_val:
                exact += 1
            else:
                abs_errors.append(abs(pred_val - gt_val))
                details.append({
                    "age": pred_row.get("age"), "col": col,
                    "pred": pred_val, "gt": gt_val,
                    "err": abs(pred_val - gt_val),
                })

    return {
        "exact_match_rate": exact / total if total else 0,
        "exact": exact, "total": total,
        "mae": sum(abs_errors) / len(abs_errors) if abs_errors else 0,
        "errors": details,
    }


# ─── Strategies ──────────────────────────────────────────────────────────

def strategy_full_single(image_path, model_name):
    """Single full-image pass with one model."""
    b64 = encode_image(image_path)
    raw = MODELS[model_name](b64, PROMPT_FULL)
    return parse_json_response(raw)


def strategy_multipass(image_path, model_name, n_passes=3):
    """Multiple passes with same model, then majority vote."""
    b64 = encode_image(image_path)
    all_passes = []
    for i in range(n_passes):
        try:
            raw = MODELS[model_name](b64, PROMPT_FULL)
            parsed = parse_json_response(raw)
            if parsed:
                all_passes.append(parsed)
                print(f"      Pass {i+1}: {len(parsed)} rows")
        except Exception as e:
            print(f"      Pass {i+1}: ERROR {e}")
    if not all_passes:
        return None
    return majority_vote(all_passes)


def strategy_cropped(image_path, model_name):
    """Crop image into sections and extract from each."""
    from PIL import Image

    # Split into top half and bottom half for focused extraction
    img = Image.open(image_path)
    w, h = img.size

    sections = [
        ("top_third", 0, 0.4),
        ("mid_third", 0.3, 0.7),
        ("bottom_third", 0.6, 1.0),
    ]

    all_rows = []
    for name, top, bottom in sections:
        cropped = img.crop((0, int(h * top), w, int(h * bottom)))
        b64 = encode_pil_image(cropped)
        try:
            raw = MODELS[model_name](b64, PROMPT_FULL)
            parsed = parse_json_response(raw)
            if parsed:
                all_rows.append(parsed)
                print(f"      {name}: {len(parsed)} rows")
        except Exception as e:
            print(f"      {name}: ERROR {e}")

    if not all_rows:
        return None

    # Merge overlapping sections via voting
    return majority_vote(all_rows)


def strategy_verify(image_path, model_name, initial_result):
    """Send initial result back with image for verification/correction."""
    if initial_result is None:
        return None
    b64 = encode_image(image_path)
    reading_str = json.dumps(initial_result, indent=2)
    prompt = PROMPT_VERIFY.format(previous_reading=reading_str)
    try:
        raw = MODELS[model_name](b64, prompt)
        return parse_json_response(raw)
    except Exception as e:
        print(f"      Verify ERROR: {e}")
        return initial_result


def preprocess_variants(image_path, preprocessing_config=None, variant_mode="full"):
    """Generate multiple enhanced variants of an image for MoE pipeline.

    Crops to relevant columns (if configured), then generates variants:
      variant_mode="full" (default): 4 variants (A, B, C, D)
        A: 2x upscale + moderate enhance (sharpen 2.0, contrast 1.5)
        B: 3x upscale + heavy enhance (sharpen 2.5, contrast 1.8)
        C: 3x upscale + binarized (threshold 140)
        D: 2x upscale + no enhance (just bigger)
      variant_mode="fast": 2 variants (A, C) — most diverse pair
        A: 2x upscale + moderate enhance (preserves grayscale)
        C: 3x upscale + binarized (pure binary — maximally different)

    Args:
        image_path: Path to the source image.
        preprocessing_config: Dict with optional "crop" key containing
            crop parameters (e.g. {"right_frac": 0.33}).
        variant_mode: "full" (4 variants) or "fast" (2 variants).

    Returns:
        List of (b64_string, variant_name) tuples.
    """
    from PIL import Image, ImageEnhance, ImageFilter

    img = Image.open(image_path).convert("RGB")
    w, h = img.size

    # Step 1: Crop to relevant columns if configured
    crop_cfg = (preprocessing_config or {}).get("crop")
    if crop_cfg:
        if crop_cfg.get("label_right_frac") is not None:
            # Composite crop: stitch [age labels] | [target columns]
            # This eliminates column confusion by hiding adjacent groups.
            label_right = int(w * crop_cfg["label_right_frac"])
            target_left = int(w * crop_cfg.get("left_frac", 0.0))
            target_right = int(w * crop_cfg.get("right_frac", 1.0))
            top = int(h * crop_cfg.get("top_frac", 0.0))
            bottom = int(h * crop_cfg.get("bottom_frac", 1.0))

            label_strip = img.crop((0, top, label_right, bottom))
            target_strip = img.crop((target_left, top, target_right, bottom))

            # Stitch side by side with a small white gap
            gap = 10
            composite_w = label_strip.width + gap + target_strip.width
            composite_h = max(label_strip.height, target_strip.height)
            img = Image.new("RGB", (composite_w, composite_h), (255, 255, 255))
            img.paste(label_strip, (0, 0))
            img.paste(target_strip, (label_strip.width + gap, 0))
            w, h = img.size
        else:
            # Simple crop (original behavior)
            left = int(w * crop_cfg.get("left_frac", 0.0))
            right = int(w * crop_cfg.get("right_frac", 1.0))
            top = int(h * crop_cfg.get("top_frac", 0.0))
            bottom = int(h * crop_cfg.get("bottom_frac", 1.0))
            img = img.crop((left, top, right, bottom))
            w, h = img.size

    variants = []

    # Variant A: 2x upscale + moderate enhance
    img_a = img.resize((w * 2, h * 2), Image.LANCZOS)
    img_a = ImageEnhance.Sharpness(img_a).enhance(2.0)
    img_a = ImageEnhance.Contrast(img_a).enhance(1.5)
    variants.append((encode_pil_image(img_a), "A_2x_moderate"))

    # Variant B: 3x upscale + heavy enhance
    # Include in fast mode too — with single-model fast mode, we need
    # 3 variants to give the ensemble enough diversity
    img_b = img.resize((w * 3, h * 3), Image.LANCZOS)
    img_b = ImageEnhance.Sharpness(img_b).enhance(2.5)
    img_b = ImageEnhance.Contrast(img_b).enhance(1.8)
    variants.append((encode_pil_image(img_b), "B_3x_heavy"))

    # Variant C: 3x upscale + binarized
    img_c = img.resize((w * 3, h * 3), Image.LANCZOS)
    img_c = img_c.convert("L").point(lambda x: 255 if x > 140 else 0, mode="1")
    img_c = img_c.convert("RGB")
    variants.append((encode_pil_image(img_c), "C_3x_binary"))

    # Variant D: 2x upscale + no enhance (just bigger)
    if variant_mode == "full":
        img_d = img.resize((w * 2, h * 2), Image.LANCZOS)
        variants.append((encode_pil_image(img_d), "D_2x_plain"))

    return variants


def strategy_moe_pipeline(image_path, preprocessing_config=None, models_to_use=None,
                          prompt_type="full", constraint_groups=None,
                          known_totals=None, custom_prompt=None,
                          variant_mode="full", return_raw_readings=False):
    """Multi-Enhancement MoE pipeline: variants × N models, constraint propagation.

    Generates enhanced image variants, sends each to all available models,
    collects all readings, applies constraint-propagation ensemble.

    Args:
        image_path: Path to the source image.
        preprocessing_config: Dict with crop/enhance config.
        models_to_use: List of model names. Auto-detected from env if None.
            When variant_mode="fast" and models_to_use is None, defaults to
            ["openai", "gemini"] (drops Claude to reduce API costs).
        prompt_type: "full" or "subtotals" (includes subtotal extraction).
        constraint_groups: Dict of subtotal_key -> [component_keys] for
            additional sum constraints.
        known_totals: Dict of age_key -> {col: value} for known fixed values
            (e.g. grand total = 1000 for per-1000 tables).
        custom_prompt: Optional custom extraction prompt string. When provided,
            overrides prompt_type selection. Used by auto-discovery pipeline.
        variant_mode: "full" (4 variants, default) or "fast" (2 variants).
        return_raw_readings: If True, return all_readings as 4th tuple element.

    Returns:
        (result_rows, resolution_log, corrections_log) — or
        (result_rows, resolution_log, corrections_log, all_readings) if
        return_raw_readings=True.
    """
    if models_to_use is None:
        env_map = {
            "openai": "OPENAI_API_KEY",
            "gemini": "GEMINI_API_KEY",
            "claude": "ANTHROPIC_API_KEY",
        }
        if variant_mode == "fast":
            # Fast mode: 2 variants × 1 model (gemini only — cheap + accurate)
            fast_models = ["gemini"]
            models_to_use = [m for m in fast_models
                             if os.environ.get(env_map[m])]
        else:
            models_to_use = [m for m in MODELS
                             if os.environ.get(env_map[m])]

    if custom_prompt is not None:
        prompt = custom_prompt
    else:
        prompt = PROMPT_FULL_WITH_SUBTOTALS if prompt_type == "subtotals" else PROMPT_FULL

    # Phase 1: Generate variants
    print("  Phase 1: Generating image variants")
    variants = preprocess_variants(image_path, preprocessing_config,
                                   variant_mode=variant_mode)
    print(f"    Generated {len(variants)} variants: {[v[1] for v in variants]}")

    # Phase 2: Send each variant to each model
    print(f"  Phase 2: Extracting from {len(variants)} variants × {len(models_to_use)} models")
    all_readings = []
    for b64, variant_name in variants:
        for model_name in models_to_use:
            label = f"{variant_name}/{model_name}"
            try:
                t0 = time.time()
                raw = MODELS[model_name](b64, prompt)
                parsed = parse_json_response(raw)
                elapsed = time.time() - t0
                if parsed:
                    all_readings.append(parsed)
                    print(f"    {label}: {len(parsed)} rows ({elapsed:.1f}s)")
                else:
                    print(f"    {label}: PARSE FAILED ({elapsed:.1f}s)")
            except Exception as e:
                print(f"    {label}: ERROR {e}")

    if not all_readings:
        print("  ERROR: No successful readings")
        if return_raw_readings:
            return None, {}, [], []
        return None, {}, []

    # Cache raw readings for diagnostic analysis
    cache_path = RESULTS_DIR / "last_readings_cache.json"
    with open(cache_path, "w") as f:
        json.dump(all_readings, f)

    print(f"  Phase 3: Constraint-propagation ensemble across {len(all_readings)} readings")
    resolved_rows, resolution_log = constraint_ensemble(
        all_readings,
        constraint_groups=constraint_groups,
        known_totals=known_totals,
    )
    print(f"    Resolved: {len(resolved_rows)} rows")

    # Report resolution methods
    methods = {}
    flagged = []
    for entry in resolution_log:
        m = entry.get("method", "unknown")
        methods[m] = methods.get(m, 0) + 1
        if entry.get("flag"):
            flagged.append(entry)

    print(f"    Resolution methods: {dict(methods)}")
    if flagged:
        print(f"    FLAGGED ({len(flagged)} cells need review):")
        for f in flagged:
            print(f"      {f.get('age', '?')}.{f.get('col', '?')}: "
                  f"{f.get('value')} [{f.get('method')}] {f.get('note', '')}")
    else:
        print("    No flagged cells — all resolved cleanly")

    # Check M+F=P consistency on final output
    constraint_ok = True
    for row in resolved_rows:
        p = row.get("persons")
        m = row.get("males")
        f = row.get("females")
        if p is not None and m is not None and f is not None:
            if p != m + f:
                constraint_ok = False
                print(f"    CONSTRAINT FAIL: {row.get('age')} "
                      f"P={p} != M={m}+F={f}={m+f}")
    if constraint_ok:
        print("    All M+F=P constraints satisfied")

    if return_raw_readings:
        return resolved_rows, resolution_log, [], all_readings
    return resolved_rows, resolution_log, []


def strategy_full_pipeline(image_path, models_to_use=None):
    """Full pipeline: multi-pass per model → ensemble → verify → constraints."""
    if models_to_use is None:
        models_to_use = [m for m in MODELS if os.environ.get({
            "openai": "OPENAI_API_KEY",
            "gemini": "GEMINI_API_KEY",
            "claude": "ANTHROPIC_API_KEY",
        }[m])]

    # Phase 1: Multi-pass per model (3 passes each)
    print("  Phase 1: Multi-pass extraction")
    model_results = {}
    for model_name in models_to_use:
        print(f"    [{model_name}]")
        result = strategy_multipass(image_path, model_name, n_passes=3)
        if result:
            model_results[model_name] = result
            print(f"      Voted: {len(result)} rows")

    if not model_results:
        return None

    # Phase 2: Cross-model ensemble
    print("  Phase 2: Cross-model ensemble")
    ensembled = cross_model_ensemble(model_results)
    print(f"    Ensembled: {len(ensembled)} rows")

    # Phase 3: Verification pass with best model
    print("  Phase 3: Verification")
    best_model = models_to_use[0]  # Use first available for verify
    verified = strategy_verify(image_path, best_model, ensembled)
    if verified:
        print(f"    Verified: {len(verified)} rows")
    else:
        verified = ensembled

    # Phase 4: Constraint enforcement
    print("  Phase 4: Constraint enforcement")
    corrected, corrections = enforce_constraints(verified)
    if corrections:
        print(f"    Made {len(corrections)} corrections:")
        for c in corrections:
            print(f"      {c.get('age', '?')} {c.get('field', '?')}: {c.get('old')} → {c.get('new')} [{c.get('rule')}]")
    else:
        print("    No corrections needed")

    return corrected


# ─── Excel Export ─────────────────────────────────────────────────────────

def export_results_to_excel(test_results, output_path=None):
    """Export pipeline results to a formatted Excel workbook with constraint verification.

    Creates one sheet per test case with:
      - Data columns: Age Group | Persons | Males | Females
      - Constraint formulas: P=M+F check per row
      - Group sum verification where applicable
      - Conditional formatting (green pass, red fail)

    Args:
        test_results: Dict of test_id -> {
            "name": str, "rows": list of row dicts,
            "constraint_groups": dict or None,
            "known_totals": dict or None,
        }
        output_path: Path for the output .xlsx. Defaults to results/extracted_data.xlsx.

    Returns:
        Path to the created Excel file.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    if output_path is None:
        output_path = RESULTS_DIR / "extracted_data.xlsx"

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Style definitions
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    check_pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    check_fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for test_id, test_data in test_results.items():
        name = test_data["name"]
        rows = test_data["rows"]
        groups = test_data.get("constraint_groups") or {}
        known = test_data.get("known_totals") or {}

        # Sheet name: truncate to 31 chars (Excel limit)
        sheet_name = name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws.merge_cells("A1:F1")
        ws["A1"] = name
        ws["A1"].font = title_font

        # Headers (row 3)
        headers = ["Age Group", "Persons", "Males", "Females", "P = M+F", "Status"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Identify subtotal and total rows for formatting
        subtotal_keys = set(groups.keys()) if groups else set()
        # Build lookup for which rows are components of which groups
        component_of = {}  # age_key -> parent subtotal key
        for sk, comps in groups.items():
            for ck in comps:
                component_of[ck] = sk

        # Write data rows starting at row 4
        data_start_row = 4
        age_to_excel_row = {}  # normalized age -> Excel row number

        for i, row in enumerate(rows):
            excel_row = data_start_row + i
            age_str = row.get("age", "")
            age_norm = normalize_age(age_str)
            age_to_excel_row[age_norm] = excel_row

            p = row.get("persons")
            m = row.get("males")
            f = row.get("females")

            ws.cell(row=excel_row, column=1, value=age_str).border = thin_border
            ws.cell(row=excel_row, column=2, value=p).border = thin_border
            ws.cell(row=excel_row, column=3, value=m).border = thin_border
            ws.cell(row=excel_row, column=4, value=f).border = thin_border

            # Number formatting with comma separator
            for col in (2, 3, 4):
                ws.cell(row=excel_row, column=col).number_format = '#,##0'

            # P=M+F check formula (column E)
            check_cell = ws.cell(row=excel_row, column=5)
            check_cell.value = f'=IF(B{excel_row}=C{excel_row}+D{excel_row},"PASS","FAIL")'
            check_cell.border = thin_border
            check_cell.alignment = Alignment(horizontal="center")

            # Status (confidence info from pipeline)
            status_parts = []
            for col_name in ("persons", "males", "females"):
                conf = row.get(f"_conf_{col_name}")
                st = row.get(f"_status_{col_name}")
                if st and st != "locked":
                    status_parts.append(f"{col_name[0].upper()}:{st}")
            status_cell = ws.cell(row=excel_row, column=6,
                                  value=", ".join(status_parts) if status_parts else "locked")
            status_cell.border = thin_border

            # Row fill for subtotals and total
            if age_norm == "total":
                for col in range(1, 7):
                    ws.cell(row=excel_row, column=col).fill = total_fill
            elif age_norm in subtotal_keys:
                for col in range(1, 7):
                    ws.cell(row=excel_row, column=col).fill = subtotal_fill

        # After data rows, add group sum verification section
        last_data_row = data_start_row + len(rows) - 1
        verify_start = last_data_row + 2

        if groups:
            ws.cell(row=verify_start, column=1, value="Constraint Verification").font = header_font
            verify_start += 1

            verify_headers = ["Constraint", "Expected (Subtotal)", "Computed (Sum)", "Check"]
            for c, h in enumerate(verify_headers, 1):
                cell = ws.cell(row=verify_start, column=c, value=h)
                cell.font = Font(bold=True)
                cell.border = thin_border

            verify_row = verify_start + 1
            for subtotal_key, comp_keys in groups.items():
                if subtotal_key not in age_to_excel_row:
                    continue

                sub_row = age_to_excel_row[subtotal_key]
                comp_rows = [age_to_excel_row[ck] for ck in comp_keys
                             if ck in age_to_excel_row]
                if not comp_rows:
                    continue

                # One verification row per column (P, M, F)
                for col_idx, col_name in [(2, "Persons"), (3, "Males"), (4, "Females")]:
                    col_letter = get_column_letter(col_idx)
                    sum_formula = "+".join(f"{col_letter}{r}" for r in comp_rows)

                    label = f"{subtotal_key} {col_name} = sum({', '.join(comp_keys)})"
                    ws.cell(row=verify_row, column=1, value=label).border = thin_border
                    ws.cell(row=verify_row, column=2,
                            value=f"={col_letter}{sub_row}").border = thin_border
                    ws.cell(row=verify_row, column=3,
                            value=f"={sum_formula}").border = thin_border
                    check = ws.cell(row=verify_row, column=4,
                                    value=f'=IF(B{verify_row}=C{verify_row},"PASS","FAIL")')
                    check.border = thin_border
                    check.alignment = Alignment(horizontal="center")
                    verify_row += 1

        # Known totals verification
        if known:
            kt_start = verify_start + len(groups) * 3 + 2 if groups else last_data_row + 2
            if not groups:
                ws.cell(row=kt_start, column=1, value="Constraint Verification").font = header_font
                kt_start += 1

            for age_key, col_vals in known.items():
                if age_key not in age_to_excel_row:
                    continue
                krow = age_to_excel_row[age_key]
                for col_name, expected in col_vals.items():
                    col_idx = {"persons": 2, "males": 3, "females": 4}.get(col_name)
                    if col_idx is None:
                        continue
                    col_letter = get_column_letter(col_idx)
                    r = kt_start
                    ws.cell(row=r, column=1,
                            value=f"Known: {age_key} {col_name} = {expected}").border = thin_border
                    ws.cell(row=r, column=2,
                            value=f"={col_letter}{krow}").border = thin_border
                    ws.cell(row=r, column=3, value=expected).border = thin_border
                    check = ws.cell(row=r, column=4,
                                    value=f'=IF(B{r}=C{r},"PASS","FAIL")')
                    check.border = thin_border
                    check.alignment = Alignment(horizontal="center")
                    kt_start += 1

        # Column widths
        ws.column_dimensions["A"].width = 18
        for col in ("B", "C", "D"):
            ws.column_dimensions[col].width = 14
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 20

        # Conditional formatting for P=M+F check column
        from openpyxl.formatting.rule import CellIsRule
        ws.conditional_formatting.add(
            f"E{data_start_row}:E{last_data_row}",
            CellIsRule(operator="equal", formula=['"PASS"'], fill=check_pass_fill))
        ws.conditional_formatting.add(
            f"E{data_start_row}:E{last_data_row}",
            CellIsRule(operator="equal", formula=['"FAIL"'], fill=check_fail_fill))

        # Conditional formatting for constraint verification checks
        if groups or known:
            max_verify_row = ws.max_row
            ws.conditional_formatting.add(
                f"D{verify_start}:D{max_verify_row}",
                CellIsRule(operator="equal", formula=['"PASS"'], fill=check_pass_fill))
            ws.conditional_formatting.add(
                f"D{verify_start}:D{max_verify_row}",
                CellIsRule(operator="equal", formula=['"FAIL"'], fill=check_fail_fill))

    wb.save(output_path)
    print(f"\nExcel exported to {output_path}")
    return output_path


def export_multigroup_to_excel(all_results, schema, output_path):
    """Export multi-group extraction results to a publication-ready Excel workbook.

    Layout mirrors the original census table:
    - One sheet per section (e.g., "ALL COMMUNITIES", "Brahmanic Hindus")
    - Columns: Age | Population P/M/F | Unmarried P/M/F | Married P/M/F | ...
    - Computed Total row at top (sum of all age rows)
    - Verification columns: M+F=P per group, cross-group sum check
    - Also creates a combined "All Sections" sheet with all data

    Args:
        all_results: Dict of section_name -> {group_name -> {"rows": [...], ...}}
        schema: TableSchema instance (for metadata).
        output_path: Path for the output .xlsx file.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Style definitions
    header_font = Font(bold=True, size=11)
    header_font_sm = Font(bold=True, size=10)
    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, size=12, italic=True)
    group_header_font = Font(bold=True, size=11, color="FFFFFF")
    total_font = Font(bold=True, size=11)
    data_font = Font(size=10)
    check_pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    check_fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    total_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    thick_bottom = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="medium"),
    )

    # Color palette for group header backgrounds
    group_colors = ["4472C4", "548235", "BF8F00", "843C0C", "7030A0"]

    # Cross-group constraint info
    cross_group = {}
    if hasattr(schema, 'cross_group_constraints'):
        cross_group = schema.cross_group_constraints or {}

    # Check if this table has a persons column
    has_persons = getattr(schema, 'has_persons_column', True)

    def _compute_total_row(rows):
        """Find the explicit TOTAL row if present, otherwise sum leaf age rows."""
        # Look for an explicit Total row in the data
        for r in rows:
            age = normalize_age(r.get("age", ""))
            if age and age.lower() == "total":
                tr = {"age": "Total"}
                for key in ("persons", "males", "females"):
                    tr[key] = r.get(key)
                return tr
        # Fallback: sum all rows (but skip subtotals and Mean Age)
        total = {"age": "Total", "persons": 0, "males": 0, "females": 0}
        for r in rows:
            total["persons"] += r.get("persons", 0) or 0
            total["males"] += r.get("males", 0) or 0
            total["females"] += r.get("females", 0) or 0
        if not has_persons:
            total["persons"] = None
        return total

    def _write_section_sheet(ws, sec_name, groups, start_row=1, include_title=True):
        """Write one section to a worksheet starting at start_row. Returns next free row."""
        if not groups:
            return start_row

        group_names = list(groups.keys())
        cur_row = start_row

        # Title
        if include_title:
            title_text = schema.title or ""
            region = schema.region or ""
            year = schema.year or ""
            if region:
                title_text = region
            if year:
                title_text = f"{title_text} — {year}"
            ws.cell(row=cur_row, column=1, value=title_text).font = title_font
            cur_row += 1

        # Section header
        ws.cell(row=cur_row, column=1, value=sec_name).font = section_font
        cur_row += 1

        # Group headers (Row: merged spans)
        header_row = cur_row
        sub_header_row = cur_row + 1
        col_offset = 2  # Column A = age labels

        group_col_map = {}
        has_xg = bool(cross_group)

        for gi, gname in enumerate(group_names):
            start_col = col_offset
            n_cols = 3  # P, M, F

            end_col = start_col + n_cols - 1
            ws.merge_cells(start_row=header_row, start_column=start_col,
                           end_row=header_row, end_column=end_col)
            cell = ws.cell(row=header_row, column=start_col, value=gname)
            cell.font = group_header_font
            cell.alignment = Alignment(horizontal="center")
            bg = group_colors[gi % len(group_colors)]
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            # Fill merged cells
            for c in range(start_col + 1, end_col + 1):
                ws.cell(row=header_row, column=c).fill = PatternFill(
                    start_color=bg, end_color=bg, fill_type="solid")

            # Sub-headers: Persons, Males, Females
            for si, sub_name in enumerate(["Persons", "Males", "Females"]):
                col = start_col + si
                cell = ws.cell(row=sub_header_row, column=col, value=sub_name)
                cell.font = header_font_sm
                cell.border = thick_bottom
                cell.alignment = Alignment(horizontal="center")

            group_col_map[gname] = start_col
            col_offset = end_col + 1

        # Cross-group check column (single column: shows "P=U+M+W+D" check)
        xg_col = None
        if has_xg:
            xg_col = col_offset
            cell = ws.cell(row=header_row, column=xg_col, value="Check")
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell = ws.cell(row=sub_header_row, column=xg_col, value="P=Sum")
            cell.font = header_font_sm
            cell.border = thick_bottom
            cell.alignment = Alignment(horizontal="center")

        # M+F=P check column
        mfp_col = (xg_col + 1) if xg_col else col_offset
        cell = ws.cell(row=header_row, column=mfp_col, value="Check")
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell = ws.cell(row=sub_header_row, column=mfp_col, value="M+F=P")
        cell.font = header_font_sm
        cell.border = thick_bottom
        cell.alignment = Alignment(horizontal="center")

        # Age header
        ws.cell(row=header_row, column=1, value="Age").font = header_font
        ws.cell(row=sub_header_row, column=1, value="").border = thick_bottom

        cur_row = sub_header_row + 1

        # Build per-group row lookup
        group_row_maps = {}
        for gname in group_names:
            rmap = {}
            for row in groups[gname].get("rows", []):
                age_norm = normalize_age(row.get("age", ""))
                if age_norm:
                    rmap[age_norm] = row
            group_row_maps[gname] = rmap

        # Determine age keys from first group (preserving order)
        first_group = group_names[0]
        age_keys = []
        for row in groups[first_group].get("rows", []):
            age_norm = normalize_age(row.get("age", ""))
            if age_norm:
                age_keys.append((age_norm, row.get("age", "")))

        # Compute total rows per group
        total_rows = {}
        for gname in group_names:
            total_rows[gname] = _compute_total_row(groups[gname].get("rows", []))

        # Write TOTAL row first (bold, highlighted)
        total_excel_row = cur_row
        ws.cell(row=cur_row, column=1, value="Total").font = total_font
        ws.cell(row=cur_row, column=1).fill = total_fill
        ws.cell(row=cur_row, column=1).border = thin_border

        for gname in group_names:
            sc = group_col_map[gname]
            tr = total_rows[gname]
            for ci, key in enumerate(["persons", "males", "females"]):
                val = tr[key]
                cell = ws.cell(row=cur_row, column=sc + ci, value=val)
                cell.font = total_font
                cell.fill = total_fill
                cell.border = thin_border
                if val is not None:
                    cell.number_format = '#,##0'

        # Total row cross-group check (formula)
        if xg_col and cross_group:
            for total_grp, comp_grps in cross_group.items():
                if total_grp not in group_col_map:
                    continue
                comps_in = [cg for cg in comp_grps if cg in group_col_map]
                if len(comps_in) < 2:
                    continue
                # Check Persons column: total_grp.P == sum(comp_grp.P)
                t_letter = get_column_letter(group_col_map[total_grp])  # Persons col
                sum_parts = [f"{get_column_letter(group_col_map[cg])}{cur_row}"
                             for cg in comps_in]
                cell = ws.cell(row=cur_row, column=xg_col)
                cell.value = f'=IF({t_letter}{cur_row}={"+".join(sum_parts)},"✓","✗")'
                cell.fill = total_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                cell.font = total_font

        # Total row M+F=P check (use first group's totals)
        first_sc = group_col_map[group_names[0]]
        p_let = get_column_letter(first_sc)
        m_let = get_column_letter(first_sc + 1)
        f_let = get_column_letter(first_sc + 2)
        cell = ws.cell(row=cur_row, column=mfp_col)
        cell.value = f'=IF({p_let}{cur_row}={m_let}{cur_row}+{f_let}{cur_row},"✓","✗")'
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        cell.font = total_font
        cur_row += 1

        # Write age-specific data rows
        data_start = cur_row
        for ri, (age_norm, age_label) in enumerate(age_keys):
            excel_row = cur_row
            is_even = (ri % 2 == 0)

            # Age label
            age_cell = ws.cell(row=excel_row, column=1, value=age_label)
            age_cell.font = data_font
            age_cell.border = thin_border
            if is_even:
                age_cell.fill = light_gray

            for gname in group_names:
                sc = group_col_map[gname]
                row_data = group_row_maps[gname].get(age_norm)

                for ci, key in enumerate(["persons", "males", "females"]):
                    val = row_data.get(key) if row_data else None
                    cell = ws.cell(row=excel_row, column=sc + ci, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.number_format = '#,##0'
                    if is_even:
                        cell.fill = light_gray

            # Cross-group check formula (Persons only — if P matches sum of components)
            if xg_col and cross_group:
                for total_grp, comp_grps in cross_group.items():
                    if total_grp not in group_col_map:
                        continue
                    comps_in = [cg for cg in comp_grps if cg in group_col_map]
                    if len(comps_in) < 2:
                        continue
                    t_letter = get_column_letter(group_col_map[total_grp])
                    sum_parts = [f"{get_column_letter(group_col_map[cg])}{excel_row}"
                                 for cg in comps_in]
                    cell = ws.cell(row=excel_row, column=xg_col)
                    cell.value = (
                        f'=IF({t_letter}{excel_row}='
                        f'{"+".join(sum_parts)},"✓",'
                        f'{t_letter}{excel_row}-({"+".join(sum_parts)}))')
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = data_font
                    if is_even:
                        cell.fill = light_gray

            # M+F=P check (all groups — show ✓ if ALL groups pass, else show which fail)
            first_sc = group_col_map[group_names[0]]
            p_let = get_column_letter(first_sc)
            m_let = get_column_letter(first_sc + 1)
            f_let = get_column_letter(first_sc + 2)
            cell = ws.cell(row=excel_row, column=mfp_col)
            cell.value = (
                f'=IF({p_let}{excel_row}={m_let}{excel_row}+'
                f'{f_let}{excel_row},"✓","✗")')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.font = data_font
            if is_even:
                cell.fill = light_gray

            cur_row += 1

        # Verification row: Total = SUM(age rows) for each P/M/F column
        verify_row = cur_row
        ws.cell(row=verify_row, column=1, value="Verify: Sum").font = Font(
            bold=True, size=9, italic=True, color="666666")
        ws.cell(row=verify_row, column=1).border = thin_border
        for gname in group_names:
            sc = group_col_map[gname]
            for ci in range(3):
                col = sc + ci
                col_letter = get_column_letter(col)
                cell = ws.cell(row=verify_row, column=col)
                cell.value = (
                    f'=IF(SUM({col_letter}{data_start}:{col_letter}{cur_row - 1})'
                    f'={col_letter}{total_excel_row},"✓",'
                    f'SUM({col_letter}{data_start}:{col_letter}{cur_row - 1})'
                    f'-{col_letter}{total_excel_row})')
                cell.font = Font(size=9, italic=True, color="666666")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
        cur_row += 1

        # Conditional formatting for check columns
        last_data_row = verify_row
        if xg_col:
            xg_letter = get_column_letter(xg_col)
            rng = f"{xg_letter}{total_excel_row}:{xg_letter}{last_data_row}"
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=['"✓"'], fill=check_pass_fill))
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="notEqual", formula=['"✓"'], fill=check_fail_fill))

        mfp_letter = get_column_letter(mfp_col)
        rng = f"{mfp_letter}{total_excel_row}:{mfp_letter}{last_data_row}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"✓"'], fill=check_pass_fill))
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="notEqual", formula=['"✓"'], fill=check_fail_fill))

        # Column widths
        ws.column_dimensions["A"].width = 14
        for gname in group_names:
            sc = group_col_map[gname]
            for c in range(sc, sc + 3):
                ws.column_dimensions[get_column_letter(c)].width = 11
        if xg_col:
            ws.column_dimensions[get_column_letter(xg_col)].width = 7
        ws.column_dimensions[get_column_letter(mfp_col)].width = 7

        return cur_row + 1  # extra blank row

    # Create per-section sheets
    for sec_name, groups in all_results.items():
        sheet_name = sec_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_section_sheet(ws, sec_name, groups)

    # Create combined "All Sections" sheet
    ws_all = wb.create_sheet(title="All Sections", index=0)
    cur_row = 1
    # Title
    title_text = schema.title or ""
    region = schema.region or ""
    if region:
        title_text = region
    if schema.year:
        title_text = f"{title_text} — {schema.year}"
    ws_all.cell(row=cur_row, column=1, value=title_text).font = Font(bold=True, size=14)
    ws_all.cell(row=cur_row + 1, column=1, value=f"Source: {schema.title}").font = Font(
        size=10, italic=True, color="666666")
    cur_row = 3

    for sec_name, groups in all_results.items():
        cur_row = _write_section_sheet(
            ws_all, sec_name, groups, start_row=cur_row, include_title=False)

    wb.save(output_path)
    print(f"\nMulti-group Excel exported to {output_path}")
    return output_path


def export_multigroup_to_json(all_results, schema, image_path, json_path):
    """Export multi-group extraction results to structured JSON.

    Args:
        all_results: Dict of section_name -> {group_name -> {"rows": [...], ...}}
        schema: TableSchema instance.
        image_path: Source image path string.
        json_path: Output JSON path.
    """
    sections_out = {}
    for sec_name, groups in all_results.items():
        sections_out[sec_name] = {}
        for group_name, gdata in groups.items():
            clean_rows = []
            for row in gdata.get("rows", []):
                clean = {k: v for k, v in row.items() if not k.startswith("_")}
                clean_rows.append(clean)
            sections_out[sec_name][group_name] = {
                "rows": clean_rows,
                "n_rows": len(clean_rows),
                "constraints_passed": gdata.get("constraints_passed", False),
                "constraint_failures": gdata.get("constraint_failures", []),
            }

    cross_group = {}
    if hasattr(schema, 'cross_group_constraints'):
        cross_group = schema.cross_group_constraints or {}

    json_output = {
        "source_image": str(image_path),
        "discovery": {
            "title": schema.title,
            "region": schema.region,
            "year": schema.year,
            "data_type": schema.data_type,
            "denominator": schema.denominator,
            "column_groups": [cg.get("name", f"Group_{i}")
                              for i, cg in enumerate(schema.column_groups)],
            "sections": [s.get("name", "All") for s in schema.sections]
                        if schema.sections else ["All"],
            "cross_group_constraints": cross_group,
        },
        "sections": sections_out,
    }

    json_path = Path(json_path)
    with open(json_path, "w") as f:
        json.dump(json_output, f, indent=2)
    print(f"  Multi-group JSON: {json_path}")
    return json_path


def run_export(test_indices=None):
    """Run MoE pipeline on test cases and export results to Excel.

    Args:
        test_indices: List of test case indices to run, or None for all.

    Returns:
        Path to the exported Excel file.
    """
    cases = TEST_CASES
    if test_indices is not None:
        cases = [TEST_CASES[i] for i in test_indices]

    available_models = []
    for m, env in [("openai", "OPENAI_API_KEY"), ("gemini", "GEMINI_API_KEY"),
                   ("claude", "ANTHROPIC_API_KEY")]:
        if os.environ.get(env):
            available_models.append(m)

    test_results = {}

    for tc in cases:
        print(f"\n{'='*70}")
        print(f"Extracting: {tc['name']}")
        print(f"{'='*70}")

        t0 = time.time()
        resolved_rows, resolution_log, _ = strategy_moe_pipeline(
            tc["image_path"],
            tc.get("preprocessing"),
            available_models,
            prompt_type=tc.get("prompt", "full"),
            constraint_groups=tc.get("constraint_groups"),
            known_totals=tc.get("known_totals"),
        )
        elapsed = time.time() - t0
        print(f"  Completed in {elapsed:.1f}s")

        # Score against GT if available
        gt = tc["ground_truth_loader"]()
        gt = apply_gt_corrections(gt, tc["id"])
        s = score(resolved_rows, gt)
        print(f"  Accuracy: {s['exact_match_rate']:.1%} ({s['exact']}/{s['total']})")
        if s["errors"]:
            for e in s["errors"]:
                print(f"    ERR {e['age']} {e['col']}: {e['pred']} vs {e['gt']} (off by {e['err']})")
        else:
            print("  PERFECT: Zero errors!")

        test_results[tc["id"]] = {
            "name": tc["name"],
            "rows": resolved_rows,
            "constraint_groups": tc.get("constraint_groups"),
            "known_totals": tc.get("known_totals"),
        }

    return export_results_to_excel(test_results)


# ─── Main ────────────────────────────────────────────────────────────────

def run_test(tc, strategy="all"):
    """Run specified strategy on a test case."""
    print(f"\n{'='*70}")
    print(f"Test: {tc['name']}")
    print(f"{'='*70}")

    gt = tc["ground_truth_loader"]()
    gt = apply_gt_corrections(gt, tc["id"])
    image_path = tc["image_path"]
    preprocessing = tc.get("preprocessing")
    results = {}

    available_models = []
    for m, env in [("openai", "OPENAI_API_KEY"), ("gemini", "GEMINI_API_KEY"), ("claude", "ANTHROPIC_API_KEY")]:
        if os.environ.get(env):
            available_models.append(m)

    if strategy in ("full", "all"):
        print("\n--- Strategy: Full Single Pass (per model) ---")
        for model_name in available_models:
            print(f"  [{model_name}]")
            try:
                t0 = time.time()
                result = strategy_full_single(image_path, model_name)
                elapsed = time.time() - t0
                s = score(result, gt)
                results[f"full_{model_name}"] = s
                print(f"    {elapsed:.1f}s | Exact: {s['exact_match_rate']:.1%} ({s['exact']}/{s['total']})")
                for e in s["errors"][:3]:
                    print(f"    ERR {e['age']} {e['col']}: {e['pred']:,} vs {e['gt']:,} (off by {e['err']:,})")
                if len(s["errors"]) > 3:
                    print(f"    ... and {len(s['errors'])-3} more errors")
            except Exception as e:
                print(f"    ERROR: {e}")

    if strategy in ("multipass", "all"):
        print("\n--- Strategy: Multi-Pass Voting (3 passes per model) ---")
        for model_name in available_models:
            print(f"  [{model_name}]")
            try:
                t0 = time.time()
                result = strategy_multipass(image_path, model_name, n_passes=3)
                elapsed = time.time() - t0
                s = score(result, gt)
                results[f"multipass_{model_name}"] = s
                print(f"    {elapsed:.1f}s | Exact: {s['exact_match_rate']:.1%} ({s['exact']}/{s['total']})")
                for e in s["errors"]:
                    print(f"    ERR {e['age']} {e['col']}: {e['pred']:,} vs {e['gt']:,} (off by {e['err']:,})")
            except Exception as e:
                print(f"    ERROR: {e}")

    if strategy in ("cropped", "all"):
        print("\n--- Strategy: Cropped Row Extraction ---")
        for model_name in available_models:
            print(f"  [{model_name}]")
            try:
                t0 = time.time()
                result = strategy_cropped(image_path, model_name)
                elapsed = time.time() - t0
                s = score(result, gt)
                results[f"cropped_{model_name}"] = s
                print(f"    {elapsed:.1f}s | Exact: {s['exact_match_rate']:.1%} ({s['exact']}/{s['total']})")
                for e in s["errors"]:
                    print(f"    ERR {e['age']} {e['col']}: {e['pred']:,} vs {e['gt']:,} (off by {e['err']:,})")
            except Exception as e:
                print(f"    ERROR: {e}")

    if strategy in ("ensemble", "all"):
        print("\n--- Strategy: Full Pipeline (multi-pass + ensemble + verify + constraints) ---")
        try:
            t0 = time.time()
            result = strategy_full_pipeline(image_path, available_models)
            elapsed = time.time() - t0
            s = score(result, gt)
            results["full_pipeline"] = s
            print(f"\n  PIPELINE RESULT: {elapsed:.1f}s | Exact: {s['exact_match_rate']:.1%} ({s['exact']}/{s['total']})")
            for e in s["errors"]:
                print(f"    ERR {e['age']} {e['col']}: {e['pred']:,} vs {e['gt']:,} (off by {e['err']:,})")
        except Exception as e:
            print(f"  PIPELINE ERROR: {e}")
            import traceback
            traceback.print_exc()

    if strategy in ("moe",):
        print("\n--- Strategy: Multi-Enhancement MoE (4 variants × N models, digit-level ensemble) ---")
        try:
            t0 = time.time()
            result, conf_meta, corrections = strategy_moe_pipeline(
                image_path, preprocessing, available_models,
                prompt_type=tc.get("prompt", "full"),
                constraint_groups=tc.get("constraint_groups"),
                known_totals=tc.get("known_totals"))
            elapsed = time.time() - t0
            s = score(result, gt)
            results["moe_pipeline"] = s
            print(f"\n  MoE RESULT: {elapsed:.1f}s | Exact: {s['exact_match_rate']:.1%} ({s['exact']}/{s['total']})")
            if s["errors"]:
                for e in s["errors"]:
                    print(f"    ERR {e['age']} {e['col']}: {e['pred']:,} vs {e['gt']:,} (off by {e['err']:,})")
            else:
                print("    PERFECT: Zero errors!")
        except Exception as e:
            print(f"  MoE PIPELINE ERROR: {e}")
            import traceback
            traceback.print_exc()

    # Summary table
    print(f"\n{'─'*50}")
    print(f"{'Strategy':<30s} {'Exact':>8s} {'Rate':>8s}")
    print(f"{'─'*50}")
    for name, s in sorted(results.items(), key=lambda x: -x[1].get("exact_match_rate", 0)):
        rate = s.get("exact_match_rate", 0)
        ex = s.get("exact", 0)
        tot = s.get("total", 0)
        print(f"  {name:<28s} {ex:>3d}/{tot:<3d}  {rate:>7.1%}")

    return results


def run_regression():
    """Run auto-discovery on all 3 test cases and compare against ground truth."""
    from schema_discovery import extract_table

    print("\n" + "=" * 70)
    print("REGRESSION TEST: Auto-Discovery Pipeline on All Test Cases")
    print("=" * 70)

    all_pass = True
    for tc in TEST_CASES:
        print(f"\n{'─'*70}")
        print(f"Test: {tc['name']}")
        print(f"{'─'*70}")

        rows, schema, config, constraints_pass, failures = extract_table(
            tc["image_path"])

        if rows is None:
            print(f"  FAIL: No rows extracted")
            all_pass = False
            continue

        # Score against GT
        gt = tc["ground_truth_loader"]()
        gt = apply_gt_corrections(gt, tc["id"])
        s = score(rows, gt)
        print(f"\n  Accuracy: {s['exact_match_rate']:.1%} ({s['exact']}/{s['total']})")
        if s["errors"]:
            all_pass = False
            for e in s["errors"]:
                print(f"    ERR {e['age']} {e['col']}: {e['pred']} vs {e['gt']} (off by {e['err']})")
        else:
            print("  PERFECT: Zero errors!")

        if not constraints_pass:
            print(f"  Constraint failures: {failures}")

    if all_pass:
        print(f"\n{'='*70}")
        print("REGRESSION: ALL TESTS PASSED")
        print(f"{'='*70}")
    else:
        print(f"\n{'='*70}")
        print("REGRESSION: SOME TESTS FAILED")
        print(f"{'='*70}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--test", type=int, default=None, help="Test case index")
    parser.add_argument("--strategy", default="all",
                        choices=["full", "multipass", "cropped", "ensemble", "moe", "all"])
    parser.add_argument("--export", action="store_true",
                        help="Run MoE pipeline and export to Excel with constraint verification")
    parser.add_argument("--extract", type=str, default=None,
                        help="Run auto-discovery pipeline on any image")
    parser.add_argument("--output", type=str, default=None,
                        help="Custom output path for --extract")
    parser.add_argument("--regression", action="store_true",
                        help="Run auto-discovery on all 3 test cases, compare against GT")
    parser.add_argument("--all-groups", action="store_true",
                        help="Extract ALL column groups (Population, Unmarried, Married, etc.) "
                             "and all sections. Use with --extract.")
    parser.add_argument("--fast", action="store_true",
                        help="Fast mode: 2 variants × 2 models instead of 4×3. "
                             "~3× faster, ~60%% cheaper. Use with --extract.")
    parser.add_argument("--oneshot", type=str, default=None,
                        help="Single-shot extraction: 1 Gemini call + constraint "
                             "verification. Usage: --oneshot IMAGE")
    parser.add_argument("--batch", type=str, default=None,
                        help="Batch single-shot extraction on all PNGs in a "
                             "directory. Usage: --batch DIR")
    parser.add_argument("--parquet", action="store_true",
                        help="Also emit Parquet output (use with --oneshot/--batch)")
    parser.add_argument("--fallback", action="store_true",
                        help="Fall back to MoE pipeline on constraint failure "
                             "(use with --oneshot)")
    args = parser.parse_args()

    if args.oneshot:
        from oneshot import extract_and_verify
        result = extract_and_verify(args.oneshot, fallback=args.fallback)
        if result and args.parquet and "df" in result:
            pq_path = RESULTS_DIR / f"{Path(args.oneshot).stem}_oneshot.parquet"
            result["df"].to_parquet(pq_path, index=False)
            print(f"Parquet: {pq_path}")
        return

    if args.batch:
        from oneshot import batch_extract
        batch_extract(args.batch, parquet=args.parquet)
        return

    if args.regression:
        run_regression()
        return

    if args.extract:
        from schema_discovery import extract_table
        output_path = args.output
        if output_path:
            output_path = Path(output_path)

        fast_mode = getattr(args, 'fast', False)

        if args.all_groups:
            all_results, schema, passed, failures = extract_table(
                args.extract, output_path=output_path,
                extract_all_groups=True, fast_mode=fast_mode)
            if passed:
                print("\nSUCCESS: All constraints verified across all groups")
            else:
                print(f"\nWARNING: {len(failures)} constraint failures")
        else:
            rows, schema, config, passed, failures = extract_table(
                args.extract, output_path=output_path,
                fast_mode=fast_mode)
            if passed:
                print("\nSUCCESS: All constraints verified")
            else:
                print(f"\nWARNING: {len(failures)} constraint failures")
        return

    if args.export:
        test_indices = [args.test] if args.test is not None else None
        run_export(test_indices)
        return

    cases = TEST_CASES
    if args.test is not None:
        cases = [TEST_CASES[args.test]]

    all_results = {}
    for tc in cases:
        result = run_test(tc, strategy=args.strategy)
        all_results[tc["id"]] = result

    # Save
    out = RESULTS_DIR / "pipeline_results.json"
    out.write_text(json.dumps(all_results, indent=2, default=str))
    print(f"\nResults saved to {out}")


if __name__ == "__main__":
    main()
