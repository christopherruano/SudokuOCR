"""Score oneshot 3-call pipeline against ground truth for all 3 test cases.

GT source: 'Ground truth for 3.xlsx' at project root.
  Sheet1: Travancore Eastern 1901 — 2 groups (Population, Unmarried), P/M/F
  Sheet2: Hyderabad State 1901 — 4 groups (Population, Unmarried, Married, Widowed), P/M/F
  Sheet3: Coorg 1901 — 3 year-groups (1901, 1891, 1881), M/F only (per 1000)
"""

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))

from pipeline import normalize_age, score
from oneshot import extract_and_verify

PROJECT_ROOT = Path(__file__).parent
GT_FILE = PROJECT_ROOT / "Ground truth for 3.xlsx"


# ---------------------------------------------------------------------------
# GT loaders — read directly from 'Ground truth for 3.xlsx'
# ---------------------------------------------------------------------------

def load_gt_sheet(sheet_name, has_persons=True):
    """Load GT from a sheet in 'Ground truth for 3.xlsx'.

    Each sheet has:
      Row 1: group headers (merged cells)
      Row 2: sub-column headers (Persons/Males/Females or Males/Females)
      Row 3+: data rows

    Returns:
        dict: {group_name: [{"age": str, "persons": int, "males": int, "females": int}, ...]}
        For M/F-only tables, persons key is omitted from the row dicts.
    """
    wb = openpyxl.load_workbook(GT_FILE, data_only=True)
    ws = wb[sheet_name]

    # Parse header row 1 to find group names and their column spans
    # Group headers are in row 1, sub-columns in row 2
    # Column A is always Age
    max_col = ws.max_column

    # Discover groups from row 1
    groups = []
    current_group = None
    current_start = None

    for c in range(2, max_col + 1):
        val = ws.cell(1, c).value
        if val is not None:
            if current_group is not None:
                groups.append({"name": str(current_group), "start": current_start, "end": c - 1})
            current_group = val
            current_start = c
    if current_group is not None:
        groups.append({"name": str(current_group), "start": current_start, "end": max_col})

    # Parse sub-columns from row 2 to determine M/F vs P/M/F per group
    for group in groups:
        sub_cols = {}
        for c in range(group["start"], group["end"] + 1):
            sub_name = str(ws.cell(2, c).value or "").strip().lower()
            if sub_name in ("persons", "males", "females"):
                sub_cols[sub_name] = c
        group["sub_cols"] = sub_cols

    # Read data rows
    result = {}
    for group in groups:
        group_name = group["name"]
        rows = []
        for r in range(3, ws.max_row + 1):
            age = ws.cell(r, 1).value
            if age is None:
                continue
            age = str(age).strip()

            row_dict = {"age": age}
            sc = group["sub_cols"]

            if "males" in sc:
                val = ws.cell(r, sc["males"]).value
                row_dict["males"] = int(val) if val is not None else None
            if "females" in sc:
                val = ws.cell(r, sc["females"]).value
                row_dict["females"] = int(val) if val is not None else None
            if "persons" in sc:
                val = ws.cell(r, sc["persons"]).value
                row_dict["persons"] = int(val) if val is not None else None

            rows.append(row_dict)
        result[group_name] = rows

    return result


def extract_predicted_for_group(parsed, group_name, section_idx=0):
    """Extract predicted rows for a specific group from oneshot parsed output.

    Args:
        parsed: Parsed oneshot output dict.
        group_name: Name of the column group to extract (case-insensitive match).
        section_idx: Which section to use (default 0 = first).

    Returns:
        list of row dicts {age, persons, males, females}
    """
    sections = parsed.get("sections", [])
    if not sections or section_idx >= len(sections):
        return []

    section = sections[section_idx]
    groups = parsed.get("metadata", {}).get("column_groups", [])

    # Find the matching group name (case-insensitive, strip trailing punctuation)
    target_group = None
    gn_clean = str(group_name).lower().strip('. ')

    # Pass 1: exact match after stripping dots/spaces
    for g in groups:
        if g.lower().strip('. ') == gn_clean:
            target_group = g
            break

    # Pass 2: year-as-int match (e.g. group_name=1901 matches "1901" or "1901.")
    if target_group is None:
        for g in groups:
            if str(group_name).strip() == str(g).strip('. '):
                target_group = g
                break

    if target_group is None:
        return []

    rows = []
    for row in section.get("rows", []):
        gdata = row.get(target_group, {})
        if not isinstance(gdata, dict):
            continue
        rows.append({
            "age": row.get("age", ""),
            "persons": gdata.get("persons"),
            "males": gdata.get("males"),
            "females": gdata.get("females"),
        })
    return rows


# ---------------------------------------------------------------------------
# Test case definitions
# ---------------------------------------------------------------------------

EVAL_CASES = [
    {
        "id": "travancore_east_1901",
        "name": "Travancore Eastern 1901",
        "image": "age_tables/Travancore/1901/Eastern_division_age_1901.png",
        "gt_sheet": "Sheet1",
        "has_persons": True,
    },
    {
        "id": "hyderabad_state_1901",
        "name": "Hyderabad State 1901",
        "image": "age_tables/Hyderabad/Hyderabad_state_summary_age_1901.png",
        "gt_sheet": "Sheet2",
        "has_persons": True,
    },
    {
        "id": "coorg_1901",
        "name": "Coorg 1901",
        "image": "age_tables/Coorg/1901/Coorg_age_1901.png",
        "gt_sheet": "Sheet3",
        "has_persons": False,
    },
]


def main():
    root = Path(__file__).parent
    results = []

    for tc in EVAL_CASES:
        print(f"\n{'#'*70}")
        print(f"# EVAL: {tc['name']}")
        print(f"{'#'*70}")

        # Run oneshot extraction
        image_path = root / tc["image"]
        result = extract_and_verify(image_path)

        if result is None or "parsed" not in result:
            print(f"\n  EXTRACTION FAILED for {tc['name']}")
            results.append({"name": tc["name"], "status": "FAILED"})
            continue

        parsed = result["parsed"]
        report = result["report"]

        # Load GT (all groups)
        gt_all = load_gt_sheet(tc["gt_sheet"], has_persons=tc["has_persons"])

        # Score each group
        total_exact = 0
        total_cells = 0
        all_errors = []
        group_scores = {}

        for gt_group_name, gt_rows in gt_all.items():
            # Find matching predicted rows
            predicted = extract_predicted_for_group(parsed, gt_group_name)

            if not predicted:
                print(f"\n  WARNING: No predicted rows for group '{gt_group_name}'")
                # Count all GT cells as missed
                for row in gt_rows:
                    for col in ("persons", "males", "females"):
                        if row.get(col) is not None:
                            total_cells += 1
                group_scores[gt_group_name] = {"exact": 0, "total": 0, "pct": "N/A"}
                continue

            sc = score(predicted, gt_rows)
            total_exact += sc["exact"]
            total_cells += sc["total"]
            all_errors.extend(
                {**e, "group": gt_group_name} for e in sc.get("errors", [])
            )
            group_scores[gt_group_name] = {
                "exact": sc["exact"],
                "total": sc["total"],
                "pct": f"{sc['exact_match_rate']:.1%}",
            }

        overall_pct = total_exact / total_cells if total_cells else 0

        # Report
        print(f"\n  GT SCORING (all groups): {total_exact}/{total_cells} = {overall_pct:.1%}")
        for gname, gs in group_scores.items():
            print(f"    {gname}: {gs['exact']}/{gs['total']} ({gs['pct']})")

        if all_errors:
            print(f"  ERRORS ({len(all_errors)}):")
            for e in all_errors[:20]:  # cap output
                print(f"    [{e.get('group', '?')}] {e['age']} {e['col']}: "
                      f"pred={e['pred']} gt={e['gt']} err={e['err']}")
            if len(all_errors) > 20:
                print(f"    ... and {len(all_errors) - 20} more")

        results.append({
            "name": tc["name"],
            "constraints": f"{report['passed']}/{report['total_checks']}",
            "all_passed": report["all_passed"],
            "gt_score": f"{total_exact}/{total_cells}",
            "gt_pct": f"{overall_pct:.1%}",
            "errors": all_errors,
            "api_calls": result.get("report", {}) and 2,
            "group_scores": group_scores,
        })

    # Summary
    print(f"\n{'='*70}")
    print("SUMMARY")
    print(f"{'='*70}")
    grand_exact = 0
    grand_total = 0
    for r in results:
        if r.get("status") == "FAILED":
            print(f"  {r['name']}: EXTRACTION FAILED")
        else:
            err_count = len(r.get("errors", []))
            status = "PERFECT" if r["gt_pct"] == "100.0%" else f"{err_count} errors"
            print(f"  {r['name']}: {r['gt_score']} ({r['gt_pct']}) "
                  f"constraints={r['constraints']} [{status}]")
            parts = r["gt_score"].split("/")
            grand_exact += int(parts[0])
            grand_total += int(parts[1])

    if grand_total:
        print(f"\n  OVERALL: {grand_exact}/{grand_total} "
              f"({grand_exact/grand_total:.1%})")


if __name__ == "__main__":
    main()
