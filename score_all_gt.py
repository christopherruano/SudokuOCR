"""Score ALL oneshot results against ALL available ground truth.

Reads saved JSON results from results/ and compares against:
1. Ground truth for 3.xlsx (multi-group GT for 3 test cases)
2. Data/*.xlsx (single-group Population GT for many provinces)
"""

import json
import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from pipeline import normalize_age, score

PROJECT_ROOT = Path(__file__).parent
RESULTS_DIR = PROJECT_ROOT / "results"
DATA_DIR = PROJECT_ROOT / "Data"
GT3_FILE = PROJECT_ROOT / "Ground truth for 3.xlsx"


# ---------------------------------------------------------------------------
# Generic GT loader from Data/*.xlsx
# ---------------------------------------------------------------------------

def fix_excel_age(val):
    """Convert Excel datetime-mangled age back to string."""
    if isinstance(val, datetime):
        return f"{val.month}-{val.day}"
    if val is None:
        return None
    return str(val).strip()


def load_gt_from_data(xlsx_name, sheet_name, district_filter=None):
    """Load GT rows from Data/*.xlsx files.

    Returns list of {age, persons, males, females}.
    """
    path = DATA_DIR / xlsx_name
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        district = row[1]
        if district is None:
            continue
        if district_filter and str(district).strip() != district_filter:
            continue
        age = fix_excel_age(row[2])
        if age is None:
            continue
        persons = row[3]
        males = row[4]
        females = row[5]
        if persons is None:
            continue
        rows.append({
            "age": age,
            "persons": int(persons),
            "males": int(float(males)) if males else 0,
            "females": int(float(females)) if females else 0,
        })
    return rows


# ---------------------------------------------------------------------------
# Multi-group GT from "Ground truth for 3.xlsx"
# ---------------------------------------------------------------------------

def load_gt3_sheet(sheet_name):
    """Load multi-group GT from 'Ground truth for 3.xlsx'."""
    wb = openpyxl.load_workbook(GT3_FILE, data_only=True)
    ws = wb[sheet_name]
    max_col = ws.max_column

    groups = []
    current_group = None
    current_start = None

    for c in range(2, max_col + 1):
        val = ws.cell(1, c).value
        if val is not None:
            if current_group is not None:
                groups.append({"name": str(current_group), "start": current_start, "end": c - 1})
            current_group = val
            current_start = c
    if current_group is not None:
        groups.append({"name": str(current_group), "start": current_start, "end": max_col})

    for group in groups:
        sub_cols = {}
        for c in range(group["start"], group["end"] + 1):
            sub_name = str(ws.cell(2, c).value or "").strip().lower()
            if sub_name in ("persons", "males", "females"):
                sub_cols[sub_name] = c
        group["sub_cols"] = sub_cols

    result = {}
    for group in groups:
        rows = []
        for r in range(3, ws.max_row + 1):
            age = ws.cell(r, 1).value
            if age is None:
                continue
            row_dict = {"age": str(age).strip()}
            sc = group["sub_cols"]
            if "males" in sc:
                val = ws.cell(r, sc["males"]).value
                row_dict["males"] = int(val) if val is not None else None
            if "females" in sc:
                val = ws.cell(r, sc["females"]).value
                row_dict["females"] = int(val) if val is not None else None
            if "persons" in sc:
                val = ws.cell(r, sc["persons"]).value
                row_dict["persons"] = int(val) if val is not None else None
            rows.append(row_dict)
        result[group["name"]] = rows

    return result


# ---------------------------------------------------------------------------
# Extract predicted rows from result JSON
# ---------------------------------------------------------------------------

def extract_predicted_group(result, group_name, section_idx=0):
    """Extract predicted rows for a specific group from result JSON."""
    data = result.get("data", [])
    if not data or section_idx >= len(data):
        return []

    section = data[section_idx]
    groups = result.get("metadata", {}).get("column_groups", [])

    target_group = None
    gn_clean = str(group_name).lower().strip(". ")

    for g in groups:
        if g.lower().strip(". ") == gn_clean:
            target_group = g
            break

    if target_group is None:
        for g in groups:
            if str(group_name).strip() == str(g).strip(". "):
                target_group = g
                break

    if target_group is None:
        return []

    rows = []
    for row in section.get("rows", []):
        gdata = row.get(target_group, {})
        if not isinstance(gdata, dict):
            continue
        rows.append({
            "age": row.get("age", ""),
            "persons": gdata.get("persons"),
            "males": gdata.get("males"),
            "females": gdata.get("females"),
        })
    return rows


def extract_predicted_population(result, section_idx=0):
    """Extract POPULATION group rows (first group) from result JSON."""
    groups = result.get("metadata", {}).get("column_groups", [])
    if not groups:
        return []
    # Try "POPULATION" first, then first group
    for name in ["POPULATION", "POPULATION.", groups[0]]:
        rows = extract_predicted_group(result, name, section_idx)
        if rows:
            return rows
    return []


# ---------------------------------------------------------------------------
# Score with flexible age matching
# ---------------------------------------------------------------------------

def score_flexible(predicted, ground_truth, mf_only=False):
    """Score predicted vs GT, handling age normalization.

    If mf_only=True, only score males and females (skip persons).
    """
    gt_by_age = {}
    for row in ground_truth:
        age_key = normalize_age(row["age"])
        if age_key:
            gt_by_age[age_key] = row

    total = 0
    exact = 0
    errors = []
    matched_keys = set()

    for pred_row in predicted:
        pred_age = normalize_age(pred_row.get("age", ""))
        if not pred_age or pred_age in matched_keys:
            continue
        if pred_age not in gt_by_age:
            continue
        gt_row = gt_by_age[pred_age]
        matched_keys.add(pred_age)

        cols = ["males", "females"] if mf_only else ["persons", "males", "females"]
        for col in cols:
            gt_val = gt_row.get(col)
            if gt_val is None:
                continue
            pred_val = pred_row.get(col)
            if pred_val is None:
                total += 1
                continue
            try:
                pred_val = int(str(pred_val).replace(",", "").replace(" ", ""))
            except (ValueError, TypeError):
                total += 1
                continue
            total += 1
            if pred_val == gt_val:
                exact += 1
            else:
                errors.append({
                    "age": pred_row.get("age"), "col": col,
                    "pred": pred_val, "gt": gt_val,
                    "err": abs(pred_val - gt_val),
                })

    unmatched = set(gt_by_age.keys()) - matched_keys
    return {
        "exact": exact, "total": total,
        "pct": exact / total if total else 0,
        "errors": errors,
        "unmatched_gt": unmatched,
        "matched": len(matched_keys),
        "gt_rows": len(gt_by_age),
    }


# ---------------------------------------------------------------------------
# Test case definitions
# ---------------------------------------------------------------------------

def load_result(filename):
    path = RESULTS_DIR / filename
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


def main():
    all_results = []

    # ===================================================================
    # Part 1: Multi-group GT from "Ground truth for 3.xlsx"
    # ===================================================================
    print("=" * 80)
    print("PART 1: Multi-group GT (Ground truth for 3.xlsx)")
    print("=" * 80)

    gt3_cases = [
        {
            "name": "Travancore Eastern 1901",
            "result_file": "Eastern_division_age_1901_oneshot.json",
            "gt_sheet": "Sheet1",
        },
        {
            "name": "Hyderabad State 1901",
            "result_file": "Hyderabad_state_summary_age_1901_oneshot.json",
            "gt_sheet": "Sheet2",
        },
        {
            "name": "Coorg 1901 (3 year groups)",
            "result_file": "Coorg_age_1901_oneshot.json",
            "gt_sheet": "Sheet3",
        },
    ]

    for tc in gt3_cases:
        result = load_result(tc["result_file"])
        if result is None:
            print(f"\n  {tc['name']}: RESULT FILE MISSING")
            all_results.append({"name": tc["name"], "exact": 0, "total": 0})
            continue

        gt_all = load_gt3_sheet(tc["gt_sheet"])
        total_exact = 0
        total_cells = 0

        print(f"\n  {tc['name']}")
        for gt_group, gt_rows in gt_all.items():
            predicted = extract_predicted_group(result, gt_group)
            if not predicted:
                print(f"    {gt_group}: NO PREDICTED ROWS")
                for row in gt_rows:
                    for col in ("persons", "males", "females"):
                        if row.get(col) is not None:
                            total_cells += 1
                continue

            # Coorg is M/F only
            mf_only = "coorg" in tc["name"].lower()
            sc = score_flexible(predicted, gt_rows, mf_only=mf_only)
            total_exact += sc["exact"]
            total_cells += sc["total"]
            status = "PERFECT" if sc["exact"] == sc["total"] else f"{len(sc['errors'])} errors"
            print(f"    {gt_group}: {sc['exact']}/{sc['total']} ({sc['pct']:.1%}) [{status}]")
            if sc["errors"]:
                for e in sc["errors"][:3]:
                    print(f"      {e['age']} {e['col']}: pred={e['pred']} gt={e['gt']} (off by {e['err']})")
                if len(sc["errors"]) > 3:
                    print(f"      ... and {len(sc['errors']) - 3} more")

        pct = total_exact / total_cells if total_cells else 0
        print(f"    TOTAL: {total_exact}/{total_cells} ({pct:.1%})")
        all_results.append({"name": tc["name"], "exact": total_exact, "total": total_cells})

    # ===================================================================
    # Part 2: Population GT from Data/*.xlsx
    # ===================================================================
    print(f"\n{'=' * 80}")
    print("PART 2: Population GT from Data/*.xlsx")
    print("=" * 80)

    # Map result files to GT sources
    data_gt_cases = [
        # Travancore
        {
            "name": "Travancore Eastern 1901 (Pop only)",
            "result_file": "Eastern_division_age_1901_oneshot.json",
            "gt_xlsx": "Travancore.xlsx",
            "gt_sheet": "Travancore_1901",
            "gt_district": "Eastern Division",
            "group": "POPULATION",
        },
        # Hyderabad State 1901
        {
            "name": "Hyderabad State 1901 (Pop only)",
            "result_file": "Hyderabad_state_summary_age_1901_oneshot.json",
            "gt_xlsx": "Hyderabad.xlsx",
            "gt_sheet": "Hyderabad_1901",
            "gt_district": "Hyderabad",
            "group": "POPULATION",
        },
        # Coorg - all 3 years from single result file
        {
            "name": "Coorg 1901 (Pop)",
            "result_file": "Coorg_age_1901_oneshot.json",
            "gt_xlsx": "Coorg.xlsx",
            "gt_sheet": "Coorg_1901",
            "gt_district": "Coorg",
            "group": "1901",
            "mf_only": True,
        },
        {
            "name": "Coorg 1891 (Pop)",
            "result_file": "Coorg_age_1901_oneshot.json",
            "gt_xlsx": "Coorg.xlsx",
            "gt_sheet": "Coorg_1891",
            "gt_district": "Coorg",
            "group": "1891",
            "mf_only": True,
        },
        {
            "name": "Coorg 1881 (Pop)",
            "result_file": "Coorg_age_1901_oneshot.json",
            "gt_xlsx": "Coorg.xlsx",
            "gt_sheet": "Coorg_1881",
            "gt_district": "Coorg",
            "group": "1881",
            "mf_only": True,
        },
    ]

    for tc in data_gt_cases:
        result = load_result(tc["result_file"])
        if result is None:
            print(f"\n  {tc['name']}: RESULT FILE MISSING")
            all_results.append({"name": tc["name"], "exact": 0, "total": 0})
            continue

        gt_rows = load_gt_from_data(tc["gt_xlsx"], tc["gt_sheet"], tc["gt_district"])
        if not gt_rows:
            print(f"\n  {tc['name']}: NO GT DATA")
            continue

        predicted = extract_predicted_group(result, tc["group"])
        if not predicted:
            # Try POPULATION
            predicted = extract_predicted_population(result)

        if not predicted:
            print(f"\n  {tc['name']}: NO PREDICTED ROWS for group '{tc['group']}'")
            all_results.append({"name": tc["name"], "exact": 0, "total": len(gt_rows) * 3})
            continue

        mf_only = tc.get("mf_only", False)
        sc = score_flexible(predicted, gt_rows, mf_only=mf_only)
        status = "PERFECT" if sc["exact"] == sc["total"] else f"{len(sc['errors'])} errors"
        print(f"\n  {tc['name']}: {sc['exact']}/{sc['total']} ({sc['pct']:.1%}) [{status}]")
        print(f"    Matched {sc['matched']}/{sc['gt_rows']} GT rows")
        if sc["unmatched_gt"]:
            print(f"    Unmatched GT ages: {sorted(sc['unmatched_gt'])}")
        if sc["errors"]:
            for e in sc["errors"][:5]:
                print(f"    {e['age']} {e['col']}: pred={e['pred']} gt={e['gt']} (off by {e['err']})")
            if len(sc["errors"]) > 5:
                print(f"    ... and {len(sc['errors']) - 5} more")
        all_results.append({"name": tc["name"], "exact": sc["exact"], "total": sc["total"]})

    # ===================================================================
    # Part 3: Hyderabad district results — constraint-only (no GT)
    # ===================================================================
    print(f"\n{'=' * 80}")
    print("PART 3: Hyderabad districts 1931/1941 (constraint checks only, no GT)")
    print("=" * 80)

    skip = {
        'Adilabad_oneshot.json', 'Cochin_age_1891_01_oneshot.json',
        'Western_division_age_1901_oneshot.json',
        'Eastern_division_age_1901_oneshot.json',
        'Hyderabad_state_summary_age_1901_oneshot.json',
        'Coorg_age_1901_oneshot.json',
        'Eastern_division_age_1901_1901_oneshot.json',
        'Coorg_age_1901_1901_oneshot.json',
    }

    constraint_pass = 0
    constraint_total = 0
    n_perfect = 0
    n_files = 0

    for f in sorted(RESULTS_DIR.glob("*_oneshot.json")):
        if f.name in skip:
            continue
        with open(f) as fp:
            data = json.load(fp)
        c = data.get("constraints", {})
        p = c.get("passed", 0)
        t = c.get("total_checks", 0)
        ok = c.get("all_passed", False)
        constraint_pass += p
        constraint_total += t
        if ok:
            n_perfect += 1
        n_files += 1
        status = "PERFECT" if ok else f"FAIL ({t-p})"
        print(f"  {f.stem:<50} {p:>6}/{t:<6} {status}")

    print(f"\n  Subtotal: {constraint_pass}/{constraint_total} "
          f"({constraint_pass/constraint_total:.1%}) — "
          f"{n_perfect}/{n_files} perfect")

    # ===================================================================
    # Grand Summary
    # ===================================================================
    print(f"\n{'=' * 80}")
    print("GRAND SUMMARY")
    print("=" * 80)

    grand_exact = sum(r["exact"] for r in all_results)
    grand_total = sum(r["total"] for r in all_results)
    print(f"\n  GT-verified cells: {grand_exact}/{grand_total} "
          f"({grand_exact/grand_total:.1%})" if grand_total else "  No GT cells")

    for r in all_results:
        pct = r["exact"] / r["total"] if r["total"] else 0
        status = "PERFECT" if r["exact"] == r["total"] else f"{r['total'] - r['exact']} wrong"
        print(f"    {r['name']:<45} {r['exact']:>4}/{r['total']:<4} ({pct:.1%}) [{status}]")

    print(f"\n  Constraint-verified (no GT): {constraint_pass}/{constraint_total} "
          f"({constraint_pass/constraint_total:.1%}) across {n_files} images")


if __name__ == "__main__":
    main()
