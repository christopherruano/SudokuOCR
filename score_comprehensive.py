"""Comprehensive GT scorer for multi-page, multi-district census results.

Handles the 1891-style format where:
- Rows = districts, Column groups = age brackets
- Multiple pages per province (each covering different age ranges)
- Pages belong to civil condition groups (Total, Unmarried, Married, Widowed)

Scores against Data/*.xlsx ground truth.
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent))
from pipeline import normalize_age

PROJECT_ROOT = Path(__file__).parent
RESULTS_DIR = PROJECT_ROOT / "results"
DATA_DIR = PROJECT_ROOT / "Data"


# ============================================================================
# GT Loading
# ============================================================================

def fix_excel_age(val):
    """Convert Excel datetime-mangled age back to string."""
    if isinstance(val, datetime):
        return f"{val.month}-{val.day}"
    if val is None:
        return None
    return str(val).strip()


def load_gt(xlsx_path, sheet_name):
    """Load GT from a Data/*.xlsx sheet.

    Returns dict: {group_name: {district: {norm_age: {persons, males, females}}}}
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]

    # Parse group headers from row 1
    # Format: each group spans 6 columns (blank, District, Age, Persons, Males, Females)
    groups = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val is not None:
            groups.append({"name": str(val).strip(), "start_col": c})

    # If no group headers, assume single "Total" group
    if not groups:
        groups = [{"name": "Total", "start_col": 1}]

    # For each group, find the District/Age/P/M/F column positions
    for i, g in enumerate(groups):
        sc = g["start_col"]
        # Check if row 2 has column headers
        row2_vals = [ws.cell(2, c).value for c in range(sc, min(sc + 7, ws.max_column + 1))]
        # Find District, Age, Persons, Males, Females positions
        g["cols"] = {}
        for offset, val in enumerate(row2_vals):
            if val is None:
                continue
            val_lower = str(val).strip().lower()
            if val_lower == "district":
                g["cols"]["district"] = sc + offset
            elif val_lower == "age":
                g["cols"]["age"] = sc + offset
            elif val_lower == "persons":
                g["cols"]["persons"] = sc + offset
            elif val_lower == "males":
                g["cols"]["males"] = sc + offset
            elif val_lower == "females":
                g["cols"]["females"] = sc + offset

        # Fallback: assume standard order if no headers found
        if not g["cols"]:
            g["cols"] = {
                "district": sc + 1,
                "age": sc + 2,
                "persons": sc + 3,
                "males": sc + 4,
                "females": sc + 5,
            }

    # Read data
    result = {}
    for g in groups:
        cols = g["cols"]
        group_data = defaultdict(dict)
        for r in range(3, ws.max_row + 1):
            district = ws.cell(r, cols.get("district", 2)).value
            if district is None:
                continue
            district = str(district).strip()
            age_raw = fix_excel_age(ws.cell(r, cols.get("age", 3)).value)
            if age_raw is None:
                continue
            age_norm = normalize_age(age_raw)
            if not age_norm:
                continue

            p = ws.cell(r, cols.get("persons", 4)).value
            m = ws.cell(r, cols.get("males", 5)).value
            f = ws.cell(r, cols.get("females", 6)).value

            try:
                p = int(float(p)) if p is not None else None
            except (ValueError, TypeError):
                p = None
            try:
                m = int(float(m)) if m is not None else None
            except (ValueError, TypeError):
                m = None
            try:
                f = int(float(f)) if f is not None else None
            except (ValueError, TypeError):
                f = None

            if p is None and m is None and f is None:
                continue

            group_data[district][age_norm] = {
                "persons": p, "males": m, "females": f
            }

        result[g["name"]] = dict(group_data)

    return result


# ============================================================================
# Result Loading & Merging
# ============================================================================

def normalize_dash(s):
    """Normalize all dash variants to hyphen."""
    return re.sub(r'[\u2010-\u2015\u2212\uFE58\uFE63\uFF0D—–]', '-', str(s))


def determine_civil_condition(metadata, column_groups):
    """Determine which civil condition group a page belongs to.

    Returns: 'Total', 'Unmarried', 'Married', or 'Widowed'
    """
    title = str(metadata.get("title", "")).lower()

    # Check column groups for "Total Unmarried" etc.
    for g in column_groups:
        g_lower = normalize_dash(g).lower()
        if "total unmarried" in g_lower or "unmarried" in g_lower:
            return "Unmarried"
        if "total married" in g_lower and "unmarried" not in g_lower:
            return "Married"
        if "total widowed" in g_lower or "widowed" in g_lower:
            return "Widowed"

    # Check title
    if "unmarried" in title:
        return "Unmarried"
    if "widowed" in title:
        return "Widowed"
    if "married" in title and "unmarried" not in title:
        return "Married"

    return "Total"


def load_page_results(province, year):
    """Load all result JSONs for a province/year.

    Returns list of page dicts with parsed data.
    """
    # Find matching result files
    prefix = province.replace(" ", "_")
    pattern = f"{prefix}_age_{year}-*_oneshot.json"
    # Also try without underscore variants
    files = sorted(RESULTS_DIR.glob(pattern))
    if not files:
        # Try case-insensitive
        pattern2 = f"{prefix.lower()}_age_{year}-*_oneshot.json"
        for f in RESULTS_DIR.glob("*_oneshot.json"):
            if f.name.lower().startswith(prefix.lower()):
                if f"_{year}-" in f.name or f"_{year}_" in f.name:
                    files.append(f)
        files = sorted(set(files))

    pages = []
    for f in files:
        with open(f) as fp:
            data = json.load(fp)
        metadata = data.get("metadata", {})
        column_groups = metadata.get("column_groups", [])
        constraints = data.get("constraints", {})

        civil_condition = determine_civil_condition(metadata, column_groups)

        # Extract age-bracket column groups (exclude "Total X" aggregates)
        age_groups = []
        for g in column_groups:
            g_norm = normalize_dash(g).lower()
            if g_norm.startswith("total ") or g_norm == "total":
                continue
            age_groups.append(g)

        # Extract district data — preserve row order for positional matching
        districts_ordered = []  # list of (name, {col_group: {P,M,F}})
        for section in data.get("data", []):
            for row in section.get("rows", []):
                # Check multiple possible name fields
                district_name = (
                    row.get("district")
                    or row.get("age", "")
                    or row.get("location", "")
                )
                if district_name:
                    district_name = str(district_name).strip()
                    if district_name.lower() in ("null", "none", ""):
                        district_name = None
                else:
                    district_name = None

                district_data = {}
                for col_group in column_groups:
                    gdata = row.get(col_group, {})
                    if not isinstance(gdata, dict):
                        continue
                    district_data[col_group] = {
                        "persons": gdata.get("persons"),
                        "males": gdata.get("males"),
                        "females": gdata.get("females"),
                    }

                districts_ordered.append((district_name, district_data))

        pages.append({
            "file": f.name,
            "civil_condition": civil_condition,
            "column_groups": column_groups,
            "age_groups": age_groups,
            "districts_ordered": districts_ordered,
            "constraints": constraints,
        })

    # Post-process: infer civil condition for ambiguous pages using sequence
    # Pages are sorted by filename (= page number), so sequential inference works
    _infer_civil_conditions(pages)

    return pages


def _infer_civil_conditions(pages):
    """Fix civil condition for pages where title was ambiguous.

    Uses age-range continuity: if a page classified as 'Total' has age ranges
    that continue from the previous page's age ranges, it inherits that cc.
    """
    AGE_ORDER = [
        "0-4", "5-9", "10-14", "15-19", "20-24", "25-29", "30-34",
        "35-39", "40-44", "45-49", "50-54", "55-59", "60andover",
    ]

    def age_rank(age_norm):
        try:
            return AGE_ORDER.index(age_norm)
        except ValueError:
            return -1

    for i, page in enumerate(pages):
        if page["civil_condition"] != "Total":
            continue
        # Check if any column group is explicitly a "Total" indicator
        has_total_marker = any(
            "total" in normalize_dash(g).lower() for g in page["column_groups"]
        )
        if has_total_marker:
            continue

        # Find the first age rank of this page
        page_ages = [normalize_age(g) for g in page["age_groups"]]
        page_ranks = [age_rank(a) for a in page_ages if age_rank(a) >= 0]
        if not page_ranks:
            continue
        first_rank = min(page_ranks)

        # Check previous page
        if i > 0:
            prev_page = pages[i - 1]
            prev_ages = [normalize_age(g) for g in prev_page["age_groups"]]
            prev_ranks = [age_rank(a) for a in prev_ages if age_rank(a) >= 0]
            if prev_ranks:
                max_prev_rank = max(prev_ranks)
                # If this page's first age follows the previous page's last age
                if first_rank == max_prev_rank + 1:
                    page["civil_condition"] = prev_page["civil_condition"]
                    continue

        # Check next page
        if i < len(pages) - 1:
            next_page = pages[i + 1]
            if next_page["civil_condition"] != "Total":
                last_rank = max(page_ranks)
                next_ages = [normalize_age(g) for g in next_page["age_groups"]]
                next_ranks = [age_rank(a) for a in next_ages if age_rank(a) >= 0]
                if next_ranks and min(next_ranks) == last_rank + 1:
                    page["civil_condition"] = next_page["civil_condition"]


def merge_pages(pages):
    """Merge multi-page results into per-district, per-group data.

    Handles pages where district names are NULL by using positional matching
    against pages that DO have names.

    Returns: {district: {gt_group: {norm_age: {persons, males, females}}}}
    """
    # Step 1: Determine district names by position
    # Find pages that have named districts
    n_rows = None
    position_names = {}  # {row_idx: district_name}

    for page in pages:
        districts_ordered = page.get("districts_ordered", [])
        if not districts_ordered:
            continue
        if n_rows is None:
            n_rows = len(districts_ordered)

        for i, (name, _) in enumerate(districts_ordered):
            if name and name.lower() not in ("null", "none", ""):
                # Clean numbered prefixes like "1 Bangalore" -> "Bangalore"
                clean = re.sub(r'^\d+\s+', '', name).strip()
                if i not in position_names:
                    position_names[i] = clean

    # Step 2: Merge data using positional matching
    merged = defaultdict(lambda: defaultdict(dict))

    for page in pages:
        cc = page["civil_condition"]
        districts_ordered = page.get("districts_ordered", [])

        for i, (raw_name, col_data) in enumerate(districts_ordered):
            # Determine district name
            district = None
            if raw_name and raw_name.lower() not in ("null", "none", ""):
                district = re.sub(r'^\d+\s+', '', raw_name).strip()
            elif i in position_names:
                district = position_names[i]
            else:
                district = f"_row_{i}"

            for col_group, values in col_data.items():
                col_norm = normalize_dash(col_group).lower()
                # Skip "Total X" columns — they're aggregates
                if col_norm.startswith("total "):
                    merged[district][cc]["_total_"] = values
                    continue
                if col_norm == "total":
                    merged[district][cc]["_total_"] = values
                    continue

                age_norm = normalize_age(col_group)
                if age_norm:
                    merged[district][cc][age_norm] = values

    # Step 3: Compute Total as sum of Unmarried + Married + Widowed where possible
    for district, groups in merged.items():
        u = groups.get("Unmarried", {})
        m = groups.get("Married", {})
        w = groups.get("Widowed", {})
        if u and m and w:
            total_data = {}
            all_ages = set(u.keys()) & set(m.keys()) & set(w.keys())
            for age in all_ages:
                if age.startswith("_"):
                    continue
                uv, mv, wv = u[age], m[age], w[age]
                total_data[age] = {}
                for col in ("persons", "males", "females"):
                    vals = [uv.get(col), mv.get(col), wv.get(col)]
                    if all(v is not None for v in vals):
                        total_data[age][col] = sum(vals)
                    else:
                        total_data[age][col] = None
            if total_data:
                # Merge into existing Total (if any), preferring computed values
                # for standard 5-year age brackets
                existing = groups.get("Total", {})
                existing.update(total_data)
                groups["Total"] = existing

    return dict(merged)


# ============================================================================
# District Name Matching
# ============================================================================

def normalize_district(name):
    """Normalize district name for matching."""
    n = str(name).strip().lower()
    # Remove common prefixes/suffixes
    n = re.sub(r'\bdistrict\b', '', n)
    n = re.sub(r'\bstate\b', '', n)
    n = re.sub(r'\bdivision\b', '', n)
    n = re.sub(r'\bplains\b', '', n)
    n = re.sub(r'\bhills\b', '', n)
    # Normalize spaces and punctuation
    n = re.sub(r'[^a-z0-9]', '', n)
    return n


def match_districts(pred_districts, gt_districts):
    """Match predicted district names to GT district names.

    Returns: dict mapping pred_name -> gt_name
    """
    mapping = {}
    gt_remaining = set(gt_districts)

    # Pass 1: exact normalized match
    for pd in pred_districts:
        pn = normalize_district(pd)
        for gd in gt_remaining:
            gn = normalize_district(gd)
            if pn == gn:
                mapping[pd] = gd
                gt_remaining.discard(gd)
                break

    # Pass 2: substring match
    unmatched_pred = [p for p in pred_districts if p not in mapping]
    for pd in unmatched_pred:
        pn = normalize_district(pd)
        best = None
        best_len = 0
        for gd in gt_remaining:
            gn = normalize_district(gd)
            # Check if one contains the other
            if pn in gn or gn in pn:
                if len(gn) > best_len:
                    best = gd
                    best_len = len(gn)
        if best:
            mapping[pd] = best
            gt_remaining.discard(best)

    return mapping


# ============================================================================
# Scoring
# ============================================================================

def score_district_group(pred_data, gt_data, mf_only=False):
    """Score one district's one group against GT.

    pred_data: {norm_age: {persons, males, females}}
    gt_data: {norm_age: {persons, males, females}}
    """
    exact = 0
    total = 0
    errors = []

    for age_norm, gt_vals in gt_data.items():
        if age_norm.startswith("_"):
            continue
        pred_vals = pred_data.get(age_norm, {})

        cols = ["males", "females"] if mf_only else ["persons", "males", "females"]
        for col in cols:
            gt_val = gt_vals.get(col)
            if gt_val is None:
                continue
            total += 1

            pred_val = pred_vals.get(col)
            if pred_val is None:
                errors.append({
                    "age": age_norm, "col": col,
                    "pred": None, "gt": gt_val, "err": gt_val,
                })
                continue

            try:
                pred_val = int(str(pred_val).replace(",", "").replace(" ", ""))
            except (ValueError, TypeError):
                errors.append({
                    "age": age_norm, "col": col,
                    "pred": pred_val, "gt": gt_val, "err": "parse",
                })
                continue

            if pred_val == gt_val:
                exact += 1
            else:
                errors.append({
                    "age": age_norm, "col": col,
                    "pred": pred_val, "gt": gt_val,
                    "err": abs(pred_val - gt_val),
                })

    return {"exact": exact, "total": total, "errors": errors}


# ============================================================================
# Province/Year Definitions
# ============================================================================

GT_MAPPING = [
    # Multi-group (1891)
    {
        "province": "Mysore", "year": "1891",
        "gt_file": "Mysore.xlsx", "gt_sheet": "Msyore_1891",
        "gt_groups": ["Total", "Unmarried", "Married", "Widowed"],
    },
    {
        "province": "Assam", "year": "1891",
        "gt_file": "Assam.xlsx", "gt_sheet": "Assam_1891",
        "gt_groups": ["Total", "Unmarried", "Married", "Widowed"],
    },
    {
        "province": "Bombay", "year": "1891",
        "gt_file": "Bombay.xlsx", "gt_sheet": "Bombay_1891",
        "gt_groups": ["Total", "Unmarried", "Married", "Widowed"],
    },
    {
        "province": "Burma", "year": "1891",
        "gt_file": "Burma.xlsx", "gt_sheet": "Burma_1891",
        "gt_groups": ["Total", "Unmarried", "Married", "Widowed"],
    },
    {
        "province": "Berar", "year": "1901",
        "gt_file": "Berar.xlsx", "gt_sheet": "Berar_1901",
        "gt_groups": ["Total", "Unmarried", "Married", "Widowed"],
    },
    # Single-group (Total only)
    {
        "province": "Madras", "year": "1891",
        "gt_file": "Madras.xlsx", "gt_sheet": "Madras_1891",
        "gt_groups": ["Total"],
    },
    {
        "province": "North_Western_Provinces_Oudh", "year": "1891",
        "gt_file": "North Western Provinces Oudh.xlsx",
        "gt_sheet": "North_Western_Provinces_Oudh_18",
        "gt_groups": ["Total"],
    },
    {
        "province": "Central_India", "year": "1891",
        "gt_file": "Central India.xlsx", "gt_sheet": "Central_India_1891",
        "gt_groups": ["Total"],
    },
    # 1901 targets
    {
        "province": "Burma", "year": "1901",
        "gt_file": "Burma.xlsx", "gt_sheet": "Burma_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Mysore", "year": "1901",
        "gt_file": "Mysore.xlsx", "gt_sheet": "Mysore_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Madras", "year": "1901",
        "gt_file": "Madras.xlsx", "gt_sheet": "Madras_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Bengal", "year": "1901",
        "gt_file": "Bengal.xlsx", "gt_sheet": "Bengal_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Central_India", "year": "1901",
        "gt_file": "Central India.xlsx", "gt_sheet": "Central_India_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Central_Provinces", "year": "1901",
        "gt_file": "Central Provinces.xlsx", "gt_sheet": "Central_Provinces_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Baroda", "year": "1901",
        "gt_file": "Baroda.xlsx", "gt_sheet": "Baroda_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Rajputana", "year": "1901",
        "gt_file": "Rajputana.xlsx", "gt_sheet": "Rajputana_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Gwalior", "year": "1901",
        "gt_file": "Gwalior.xlsx", "gt_sheet": "Central_India_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Bombay", "year": "1901",
        "gt_file": "Bombay.xlsx", "gt_sheet": "Bombay_1891",  # check
        "gt_groups": ["Total"],
    },
    {
        "province": "Ajmer_Merwara", "year": "1901",
        "gt_file": "Ajmer Merwara.xlsx", "gt_sheet": "Ajmer_Merwara_1901",
        "gt_groups": ["Total"],
    },
    {
        "province": "Assam", "year": "1901",
        "gt_file": "Assam.xlsx", "gt_sheet": "Assam_1901",
        "gt_groups": ["Total"],
    },
]


# ============================================================================
# Main
# ============================================================================

def score_province(config, verbose=True):
    """Score one province/year against GT."""
    province = config["province"]
    year = config["year"]
    gt_file = DATA_DIR / config["gt_file"]
    gt_sheet = config["gt_sheet"]
    gt_group_names = config["gt_groups"]

    if not gt_file.exists():
        return {"status": "GT_MISSING", "province": province, "year": year}

    # Load GT
    gt_all = load_gt(gt_file, gt_sheet)
    if not gt_all:
        return {"status": "GT_EMPTY", "province": province, "year": year}

    # Load pipeline results
    pages = load_page_results(province, year)
    if not pages:
        return {"status": "NO_RESULTS", "province": province, "year": year}

    # Merge pages
    merged = merge_pages(pages)

    # Count constraints across all pages
    total_constraint_pass = sum(p["constraints"].get("passed", 0) for p in pages)
    total_constraint_checks = sum(p["constraints"].get("total_checks", 0) for p in pages)

    # Score each GT group
    grand_exact = 0
    grand_total = 0
    all_errors = []
    group_results = {}

    for gt_group_name in gt_group_names:
        gt_group_data = gt_all.get(gt_group_name, {})
        if not gt_group_data:
            if verbose:
                print(f"    WARNING: No GT data for group '{gt_group_name}'")
            continue

        gt_districts = list(gt_group_data.keys())
        pred_districts = list(merged.keys())

        # Match districts
        district_map = match_districts(pred_districts, gt_districts)

        group_exact = 0
        group_total = 0
        group_errors = []
        matched_districts = 0

        for pred_name, gt_name in district_map.items():
            # Get predicted data for this group
            pred_group_data = merged.get(pred_name, {}).get(gt_group_name, {})
            if not pred_group_data:
                # Try case-insensitive group match
                for pg_name in merged.get(pred_name, {}):
                    if pg_name.lower() == gt_group_name.lower():
                        pred_group_data = merged[pred_name][pg_name]
                        break
            if not pred_group_data:
                continue

            gt_district_data = gt_group_data[gt_name]
            matched_districts += 1

            sc = score_district_group(pred_group_data, gt_district_data)
            group_exact += sc["exact"]
            group_total += sc["total"]
            for e in sc["errors"]:
                e["district"] = gt_name
                e["group"] = gt_group_name
            group_errors.extend(sc["errors"])

        grand_exact += group_exact
        grand_total += group_total
        all_errors.extend(group_errors)

        pct = group_exact / group_total if group_total else 0
        group_results[gt_group_name] = {
            "exact": group_exact, "total": group_total, "pct": pct,
            "errors": len(group_errors),
            "districts_matched": matched_districts,
            "districts_gt": len(gt_districts),
        }

        if verbose:
            status = "PERFECT" if group_exact == group_total else f"{len(group_errors)} errors"
            print(f"    {gt_group_name}: {group_exact}/{group_total} ({pct:.1%}) "
                  f"[{matched_districts}/{len(gt_districts)} districts] [{status}]")
            if group_errors and verbose:
                for e in group_errors[:3]:
                    print(f"      {e['district']}/{e['age']} {e['col']}: "
                          f"pred={e['pred']} gt={e['gt']} (err={e['err']})")
                if len(group_errors) > 3:
                    print(f"      ... and {len(group_errors) - 3} more")

    grand_pct = grand_exact / grand_total if grand_total else 0

    return {
        "status": "OK",
        "province": province,
        "year": year,
        "pages": len(pages),
        "exact": grand_exact,
        "total": grand_total,
        "pct": grand_pct,
        "errors": all_errors,
        "groups": group_results,
        "constraint_pass": total_constraint_pass,
        "constraint_total": total_constraint_checks,
    }


def main():
    print("=" * 80)
    print("COMPREHENSIVE GT SCORING: Multi-page Census Results vs Data/*.xlsx")
    print("=" * 80)

    all_results = []
    grand_exact = 0
    grand_total = 0
    grand_constraint_pass = 0
    grand_constraint_total = 0
    provinces_with_results = 0
    provinces_perfect = 0

    for config in GT_MAPPING:
        province = config["province"]
        year = config["year"]
        print(f"\n{'─' * 60}")
        print(f"  {province} {year}")
        print(f"{'─' * 60}")

        result = score_province(config)
        all_results.append(result)

        if result["status"] != "OK":
            print(f"    {result['status']}")
            continue

        provinces_with_results += 1
        grand_exact += result["exact"]
        grand_total += result["total"]
        grand_constraint_pass += result["constraint_pass"]
        grand_constraint_total += result["constraint_total"]

        pct = result["pct"]
        is_perfect = result["exact"] == result["total"]
        if is_perfect:
            provinces_perfect += 1

        c_pct = result["constraint_pass"] / result["constraint_total"] if result["constraint_total"] else 0
        print(f"    OVERALL: {result['exact']}/{result['total']} ({pct:.1%}) "
              f"from {result['pages']} pages")
        print(f"    Constraints: {result['constraint_pass']}/{result['constraint_total']} ({c_pct:.1%})")

    # Grand summary
    print(f"\n{'=' * 80}")
    print("GRAND SUMMARY")
    print(f"{'=' * 80}")

    grand_pct = grand_exact / grand_total if grand_total else 0
    grand_c_pct = grand_constraint_pass / grand_constraint_total if grand_constraint_total else 0

    print(f"\n  Provinces scored: {provinces_with_results} ({provinces_perfect} perfect)")
    print(f"  GT cells: {grand_exact}/{grand_total} ({grand_pct:.1%})")
    print(f"  Constraints: {grand_constraint_pass}/{grand_constraint_total} ({grand_c_pct:.1%})")

    print(f"\n  {'Province':<35} {'Pages':>5} {'GT Score':>12} {'Rate':>8} {'Constraints':>15}")
    print(f"  {'─' * 80}")
    for r in all_results:
        if r["status"] != "OK":
            print(f"  {r['province']+' '+r['year']:<35} {'—':>5} {r['status']:>12}")
            continue
        pct = f"{r['pct']:.1%}"
        c_str = f"{r['constraint_pass']}/{r['constraint_total']}"
        status = "PERFECT" if r["exact"] == r["total"] else f"{len(r['errors'])} err"
        print(f"  {r['province']+' '+r['year']:<35} {r['pages']:>5} "
              f"{r['exact']:>5}/{r['total']:<5} {pct:>8} {c_str:>15}  [{status}]")

    # Also score constraint-only results (Hyderabad districts etc.)
    print(f"\n{'=' * 80}")
    print("CONSTRAINT-ONLY RESULTS (no GT match)")
    print(f"{'=' * 80}")

    scored_files = set()
    for r in all_results:
        if r["status"] == "OK":
            pages = load_page_results(r["province"], r["year"])
            for p in pages:
                scored_files.add(p["file"])

    c_only_pass = 0
    c_only_total = 0
    c_only_count = 0
    for f in sorted(RESULTS_DIR.glob("*_oneshot.json")):
        if f.name in scored_files:
            continue
        with open(f) as fp:
            data = json.load(fp)
        c = data.get("constraints", {})
        p = c.get("passed", 0)
        t = c.get("total_checks", 0)
        if t == 0:
            continue
        c_only_pass += p
        c_only_total += t
        c_only_count += 1
        ok = "PASS" if c.get("all_passed") else f"FAIL({t - p})"
        print(f"  {f.stem:<55} {p:>5}/{t:<5} {ok}")

    if c_only_total:
        print(f"\n  Constraint-only: {c_only_pass}/{c_only_total} "
              f"({c_only_pass / c_only_total:.1%}) across {c_only_count} files")

    # Combined totals
    print(f"\n{'=' * 80}")
    print("COMBINED TOTALS")
    print(f"{'=' * 80}")
    all_c_pass = grand_constraint_pass + c_only_pass
    all_c_total = grand_constraint_total + c_only_total
    all_c_pct = all_c_pass / all_c_total if all_c_total else 0
    total_files = provinces_with_results + c_only_count  # approximate

    print(f"  GT-verified: {grand_exact}/{grand_total} ({grand_pct:.1%})")
    print(f"  All constraints: {all_c_pass}/{all_c_total} ({all_c_pct:.1%})")
    print(f"  Total result files scored: {len(scored_files) + c_only_count}")


if __name__ == "__main__":
    main()
