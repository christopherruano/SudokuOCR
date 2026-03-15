"""
Baseline evaluation: test OpenAI, Gemini, and Claude vision APIs
on historical Indian census table OCR.

Sends census table images to each API, asks for structured number extraction,
then compares against hand-transcribed ground truth from Excel files.

Usage:
    python3 baseline_eval.py                  # run all test cases
    python3 baseline_eval.py --test 0         # run specific test case
    python3 baseline_eval.py --api openai     # run specific API only
"""

import os
import sys
import json
import base64
import time
import subprocess
import argparse
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

load_dotenv()

PROJECT_ROOT = Path(__file__).parent
AGE_TABLES = PROJECT_ROOT / "age_tables"
DATA_DIR = PROJECT_ROOT / "Data"
RESULTS_DIR = PROJECT_ROOT / "results"
RESULTS_DIR.mkdir(exist_ok=True)


# ─── Utilities ───────────────────────────────────────────────────────────

def fix_excel_age(age):
    """Excel sometimes interprets age ranges like '5-10' as dates."""
    if hasattr(age, 'strftime'):
        return f"{age.month}-{age.day}"
    return str(age)


def normalize_age(age_str):
    """Normalize age string for matching."""
    s = str(age_str).replace("\u2013", "-").replace(" ", "").lower()
    s = s.replace("total", "").strip()
    return s


def pdf_to_images(pdf_path, output_dir=None, dpi=300):
    """Convert PDF pages to PNG images using pdftoppm."""
    pdf_path = Path(pdf_path)
    if output_dir is None:
        output_dir = pdf_path.parent / f"{pdf_path.stem}_pages"
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    prefix = output_dir / "page"
    subprocess.run(
        ["pdftoppm", "-png", "-r", str(dpi), str(pdf_path), str(prefix)],
        check=True, capture_output=True,
    )
    return sorted(output_dir.glob("*.png"))


def encode_image(image_path):
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def parse_json_response(text):
    """Extract JSON from a response that might have markdown fences or prose."""
    import re
    text = text.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        text = "\n".join(lines)
    # Remove Python-style underscore number separators (e.g. 1_355_867 -> 1355867)
    text = re.sub(r'(?<=\d)_(?=\d)', '', text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start = text.find("[")
        end = text.rfind("]") + 1
        if start >= 0 and end > start:
            try:
                return json.loads(text[start:end])
            except json.JSONDecodeError:
                pass
    return None


# ─── Ground Truth Loaders ────────────────────────────────────────────────

def load_gt_simple(xlsx_name, sheet_name, district_filter=None):
    """Generic loader for xlsx files with standard 6-col format:
    [None, District, Age, Persons, Males, Females]
    Returns list of dicts with keys: age, persons, males, females.
    """
    wb = openpyxl.load_workbook(DATA_DIR / xlsx_name, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        district = row[1]
        if district is None:
            continue
        if district_filter and district != district_filter:
            continue
        age = fix_excel_age(row[2])
        persons = row[3]
        males = row[4]
        females = row[5]
        if persons is None:
            continue
        rows.append({
            "age": age,
            "persons": int(persons),
            "males": int(float(males)) if males else 0,
            "females": int(float(females)) if females else 0,
        })
    return rows


def load_gt_berar_full():
    """Berar has a wider format: Total + Unmarried + Married + Widowed sections."""
    wb = openpyxl.load_workbook(DATA_DIR / "Berar.xlsx", data_only=True)
    ws = wb["Berar_1901"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        # Unmarried section: cols H(7), I(8), J(9), K(10), L(11)
        district = row[7]
        if district is None:
            continue
        age = fix_excel_age(row[8])
        rows.append({
            "district": str(district),
            "age": age,
            "unmarried_males": int(row[10]) if row[10] else 0,
            "unmarried_females": int(row[11]) if row[11] else 0,
            # Married section: cols N(13), O(14), P(15), Q(16), R(17)
            "married_males": int(row[16]) if row[16] else 0,
            "married_females": int(row[17]) if row[17] else 0,
            # Widowed section: cols T(19), U(20), V(21), W(22), X(23)
            "widowed_males": int(row[22]) if row[22] else 0,
            "widowed_females": int(row[23]) if row[23] else 0,
        })
    return rows


# ─── Comparison / Scoring ────────────────────────────────────────────────

def compare_population_table(predicted, ground_truth, num_cols=("persons", "males", "females")):
    """Generic comparison for population tables.

    Matches predicted rows to ground truth by age group,
    then compares numeric columns.
    """
    if predicted is None:
        return {"error": "Could not parse response as JSON", "exact_match_rate": 0}

    gt_by_age = {}
    for row in ground_truth:
        key = normalize_age(row["age"])
        gt_by_age[key] = row

    results = []
    total_cells = 0
    exact_matches = 0
    abs_errors = []
    rel_errors = []
    matched_gt_keys = set()

    for pred_row in predicted:
        pred_age = normalize_age(pred_row.get("age", ""))

        # Find matching ground truth row
        gt_row = gt_by_age.get(pred_age)
        if gt_row is None:
            # Fuzzy: check containment
            for key in gt_by_age:
                if key not in matched_gt_keys and (key in pred_age or pred_age in key):
                    gt_row = gt_by_age[key]
                    break

        if gt_row is None:
            results.append({"age": pred_row.get("age"), "status": "no_gt_match"})
            continue

        matched_gt_keys.add(normalize_age(gt_row["age"]))
        row_result = {"age": pred_row.get("age"), "cells": {}}

        for col in num_cols:
            pred_val = pred_row.get(col)
            gt_val = gt_row.get(col)
            if gt_val is None:
                continue

            if pred_val is None:
                row_result["cells"][col] = {"pred": None, "gt": gt_val, "exact": False}
                total_cells += 1
                abs_errors.append(abs(gt_val))
                rel_errors.append(1.0)
                continue

            try:
                pred_val = int(str(pred_val).replace(",", "").replace(" ", ""))
            except (ValueError, TypeError):
                row_result["cells"][col] = {"pred": pred_val, "gt": gt_val, "exact": False}
                total_cells += 1
                abs_errors.append(abs(gt_val))
                rel_errors.append(1.0)
                continue

            total_cells += 1
            is_exact = pred_val == gt_val
            if is_exact:
                exact_matches += 1
            abs_err = abs(pred_val - gt_val)
            rel_err = abs_err / abs(gt_val) if gt_val != 0 else (0 if pred_val == 0 else 1)
            abs_errors.append(abs_err)
            rel_errors.append(rel_err)
            row_result["cells"][col] = {
                "pred": pred_val, "gt": gt_val,
                "exact": is_exact, "abs_err": abs_err,
                "rel_err": round(rel_err, 6),
            }
        results.append(row_result)

    # Check for unmatched ground truth rows
    unmatched = [r for r in ground_truth if normalize_age(r["age"]) not in matched_gt_keys]

    return {
        "exact_match_rate": exact_matches / total_cells if total_cells else 0,
        "total_cells": total_cells,
        "exact_matches": exact_matches,
        "mean_absolute_error": sum(abs_errors) / len(abs_errors) if abs_errors else 0,
        "mean_relative_error": sum(rel_errors) / len(rel_errors) if rel_errors else 0,
        "unmatched_gt_rows": len(unmatched),
        "per_row": results,
    }


# ─── API Callers ─────────────────────────────────────────────────────────

def call_openai(image_path, prompt):
    from openai import OpenAI
    client = OpenAI()
    b64 = encode_image(image_path)
    ext = Path(image_path).suffix.lower()
    mime = "image/png" if ext == ".png" else "image/jpeg"

    response = client.chat.completions.create(
        model="gpt-4.1",
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {
                    "url": f"data:{mime};base64,{b64}",
                    "detail": "high"
                }},
            ],
        }],
        max_tokens=8192,
        temperature=0,
    )
    return response.choices[0].message.content


def call_gemini(image_path, prompt):
    from google import genai
    client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])

    b64 = encode_image(image_path)
    ext = Path(image_path).suffix.lower()
    mime = "image/png" if ext == ".png" else "image/jpeg"

    response = client.models.generate_content(
        model="gemini-2.5-pro",
        contents=[
            prompt,
            genai.types.Part.from_bytes(data=base64.b64decode(b64), mime_type=mime),
        ],
        config=genai.types.GenerateContentConfig(temperature=0),
    )
    return response.text


def call_claude(image_path, prompt):
    import anthropic
    client = anthropic.Anthropic()
    b64 = encode_image(image_path)
    ext = Path(image_path).suffix.lower()
    media_type = "image/png" if ext == ".png" else "image/jpeg"

    response = client.messages.create(
        model="claude-opus-4-20250514",
        max_tokens=8192,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": b64,
                }},
                {"type": "text", "text": prompt},
            ],
        }],
    )
    return response.content[0].text


APIS = {
    "openai": {"caller": call_openai, "env_key": "OPENAI_API_KEY"},
    "gemini": {"caller": call_gemini, "env_key": "GEMINI_API_KEY"},
    "claude": {"caller": call_claude, "env_key": "ANTHROPIC_API_KEY"},
}


# ─── Prompts ─────────────────────────────────────────────────────────────

PROMPT_POPULATION_SUMMARY = """This is a scanned historical Indian census table showing age, sex and civil condition data.

Extract the POPULATION totals (columns: Persons, Males, Females) for each SUMMARY age group row.
Summary age groups are: "Total 0-5" (or just "0-5"), "5-10", "10-15", "15-20", "20-25", "25-30", "30-35", "35-40", "40-45", "45-50", "50-55", "55-60", "60 and over", and optionally "Total".

Do NOT include individual year rows (0-1, 1-2, etc.) — only the grouped totals.
If a "Total" or grand total row exists at the bottom, include it.

Return ONLY a JSON array of objects with keys: "age", "persons", "males", "females".
Every value for persons/males/females must be an integer (no commas).
Return ONLY valid JSON, no other text."""

PROMPT_FULL_TABLE = """This is a scanned historical Indian census table.

Extract ALL numeric data from this table. Return as a JSON array of objects.
Each object should represent one row and have keys matching the column headers.
For the age column, use "age". For numeric columns use descriptive keys like
"population_persons", "population_males", "population_females",
"unmarried_persons", "unmarried_males", "unmarried_females",
"married_persons", "married_males", "married_females",
"widowed_persons", "widowed_males", "widowed_females".

Only include summary age group rows (0-5, 5-10, etc.), not individual years.
All numeric values must be integers (no commas).
Return ONLY valid JSON, no other text."""


# ─── Test Cases ──────────────────────────────────────────────────────────

TEST_CASES = [
    {
        "id": "travancore_east_1901",
        "name": "Travancore Eastern Division 1901 (PNG, clean)",
        "image_path": AGE_TABLES / "Travancore/1901/Eastern_division_age_1901.png",
        "ground_truth_loader": lambda: load_gt_simple(
            "Travancore.xlsx", "Travancore_1901", "Eastern Division"),
        "prompt": PROMPT_POPULATION_SUMMARY,
    },
    {
        "id": "hyderabad_state_1901",
        "name": "Hyderabad State Summary 1901 (PNG, dense)",
        "image_path": AGE_TABLES / "Hyderabad/Hyderabad_state_summary_age_1901.png",
        "ground_truth_loader": lambda: load_gt_simple(
            "Hyderabad.xlsx", "Hyderabad_1901", "Hyderabad"),
        "prompt": PROMPT_POPULATION_SUMMARY,
    },
    {
        "id": "coorg_1901",
        "name": "Coorg 1901 (PNG, per-1000 proportions)",
        "image_path": AGE_TABLES / "Coorg/1901/Coorg_age_1901.png",
        "ground_truth_loader": lambda: load_gt_simple(
            "Coorg.xlsx", "Coorg_1901", "Coorg"),
        "prompt": """This is a scanned historical Indian census table showing age distribution per 1,000 of each sex for multiple census years.

Extract the data for the 1901 census (the leftmost data columns after the age column).
The columns are: Males, Females.

Return ONLY a JSON array of objects with keys: "age", "persons" (sum of males+females if a persons column exists, otherwise null), "males", "females".
All numeric values must be integers.
Only include individual age rows and totals visible in the table.
Return ONLY valid JSON, no other text.""",
    },
]


# ─── Main ────────────────────────────────────────────────────────────────

def run_single_test(tc, api_filter=None):
    """Run a single test case across all available APIs."""
    print(f"\n{'='*70}")
    print(f"Test: {tc['name']}")
    print(f"Image: {tc['image_path']}")
    print(f"{'='*70}")

    if not tc["image_path"].exists():
        print(f"  IMAGE NOT FOUND: {tc['image_path']}")
        return {}

    ground_truth = tc["ground_truth_loader"]()
    if ground_truth:
        print(f"Ground truth: {len(ground_truth)} rows")
        for r in ground_truth[:5]:
            print(f"  {r['age']:>15s}: P={r['persons']:>10,}  M={r['males']:>10,}  F={r['females']:>10,}")
        if len(ground_truth) > 5:
            print(f"  ... and {len(ground_truth) - 5} more rows")
    else:
        print("  No ground truth available (qualitative test only)")

    test_results = {}

    for api_name, api_info in APIS.items():
        if api_filter and api_name != api_filter:
            continue

        env_key = api_info["env_key"]
        if not os.environ.get(env_key):
            print(f"\n  [{api_name.upper()}] SKIPPED (no {env_key})")
            continue

        print(f"\n  [{api_name.upper()}]")
        try:
            t0 = time.time()
            raw = api_info["caller"](tc["image_path"], tc["prompt"])
            elapsed = time.time() - t0
            print(f"    Time: {elapsed:.1f}s")

            # Save raw response
            out_prefix = f"{tc['id']}_{api_name}"
            (RESULTS_DIR / f"{out_prefix}_raw.txt").write_text(raw)

            parsed = parse_json_response(raw)
            if parsed:
                print(f"    Parsed: {len(parsed)} rows")
                (RESULTS_DIR / f"{out_prefix}_parsed.json").write_text(
                    json.dumps(parsed, indent=2))
            else:
                print(f"    FAILED to parse JSON")
                print(f"    Response preview: {raw[:300]}")

            if ground_truth and parsed:
                comp = compare_population_table(parsed, ground_truth)
                test_results[api_name] = comp
                em = comp["exact_match_rate"]
                print(f"    Exact match: {em:.1%} ({comp['exact_matches']}/{comp['total_cells']} cells)")
                print(f"    MAE: {comp['mean_absolute_error']:,.1f}")
                print(f"    MRE: {comp['mean_relative_error']:.4%}")
                if comp.get("unmatched_gt_rows"):
                    print(f"    WARNING: {comp['unmatched_gt_rows']} ground truth rows not matched")

                # Show errors
                for row in comp.get("per_row", []):
                    cells = row.get("cells", {})
                    errors = {k: v for k, v in cells.items()
                              if isinstance(v, dict) and not v.get("exact", True)}
                    if errors:
                        print(f"    Age {row['age']}:")
                        for col, d in errors.items():
                            pred = d['pred']
                            gt = d['gt']
                            err = d.get('abs_err', '?')
                            pred_s = f"{pred:>10,}" if isinstance(pred, (int, float)) else f"{pred!r:>10s}"
                        gt_s = f"{gt:>10,}" if isinstance(gt, (int, float)) else f"{gt!r:>10s}"
                        err_s = f"{err:>8,}" if isinstance(err, (int, float)) else f"{err!r:>8s}"
                        print(f"      {col}: pred={pred_s} gt={gt_s} err={err_s}")

                # Save comparison
                (RESULTS_DIR / f"{out_prefix}_comparison.json").write_text(
                    json.dumps(comp, indent=2, default=str))
            elif not ground_truth:
                test_results[api_name] = {"note": "no ground truth", "parsed_rows": len(parsed) if parsed else 0}
                if parsed:
                    print(f"    (No ground truth to compare — saved {len(parsed)} parsed rows)")

        except Exception as e:
            print(f"    ERROR: {e}")
            import traceback
            traceback.print_exc()
            test_results[api_name] = {"error": str(e)}

    return test_results


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--test", type=int, help="Run specific test case by index")
    parser.add_argument("--api", type=str, help="Run specific API only (openai/gemini/claude)")
    args = parser.parse_args()

    cases = TEST_CASES
    if args.test is not None:
        cases = [TEST_CASES[args.test]]

    all_results = {}
    for tc in cases:
        result = run_single_test(tc, api_filter=args.api)
        all_results[tc["id"]] = result

    # Overall summary
    print(f"\n{'='*70}")
    print("OVERALL SUMMARY")
    print(f"{'='*70}")

    # Collect per-API aggregates
    api_totals = {}
    for test_id, test_result in all_results.items():
        for api_name, comp in test_result.items():
            if "error" in comp or "note" in comp:
                continue
            if api_name not in api_totals:
                api_totals[api_name] = {"exact": 0, "total": 0, "abs_errors": [], "rel_errors": []}
            api_totals[api_name]["exact"] += comp["exact_matches"]
            api_totals[api_name]["total"] += comp["total_cells"]

    for api_name, totals in api_totals.items():
        rate = totals["exact"] / totals["total"] if totals["total"] else 0
        print(f"  {api_name:>8s}: {rate:.1%} overall exact match "
              f"({totals['exact']}/{totals['total']} cells across all tests)")

    # Per-test breakdown
    print()
    for test_id, test_result in all_results.items():
        print(f"  {test_id}:")
        for api_name, comp in test_result.items():
            if "error" in comp:
                print(f"    {api_name:>8s}: ERROR")
            elif "note" in comp:
                print(f"    {api_name:>8s}: {comp['note']}")
            else:
                print(f"    {api_name:>8s}: {comp['exact_match_rate']:.1%} "
                      f"({comp['exact_matches']}/{comp['total_cells']})")

    # Save all
    (RESULTS_DIR / "all_results.json").write_text(
        json.dumps(all_results, indent=2, default=str))
    print(f"\nResults saved to {RESULTS_DIR}/")


if __name__ == "__main__":
    main()
